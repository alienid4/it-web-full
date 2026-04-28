"""
recon_service.py - Excel ↔ DB 對帳 (P4 of CMDB)
功能: 上傳 .xlsx / .csv, 抓 hostname + IP 欄位, 跟 hosts collection 比對
"""
import csv
import io
from services.mongo_service import get_collection


# 可能的 hostname 欄位名 (大小寫不敏感)
HOSTNAME_KEYS = {"hostname", "host", "host_name", "主機", "主機名稱", "電腦名稱", "name"}
IP_KEYS = {"ip", "ip_address", "ipaddress", "ip 位址", "ip位址", "管理ip"}


def _norm(s):
    return str(s or "").strip().lower().replace(" ", "")


def _find_columns(header):
    """找 hostname / ip 在 header (list) 的 index"""
    h_idx = ip_idx = None
    for i, col in enumerate(header):
        nc = _norm(col)
        if h_idx is None and nc in HOSTNAME_KEYS:
            h_idx = i
        if ip_idx is None and nc in IP_KEYS:
            ip_idx = i
    return h_idx, ip_idx


def parse_csv(file_bytes):
    """parse CSV bytes → list of {hostname, ip, raw}"""
    # 嘗試 UTF-8, BIG5, GBK
    text = None
    for enc in ("utf-8-sig", "utf-8", "big5", "gbk"):
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("無法解碼 CSV (試了 utf-8/big5/gbk 都失敗)")

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    header = rows[0]
    h_idx, ip_idx = _find_columns(header)
    if h_idx is None and ip_idx is None:
        raise ValueError(f"找不到 hostname / ip 欄位 (header: {header})")

    out = []
    for r in rows[1:]:
        if not r or all(not c.strip() for c in r):
            continue
        hostname = r[h_idx].strip() if h_idx is not None and len(r) > h_idx else ""
        ip = r[ip_idx].strip() if ip_idx is not None and len(r) > ip_idx else ""
        if hostname or ip:
            out.append({"hostname": hostname, "ip": ip, "raw": dict(zip(header, r))})
    return out


def parse_xlsx(file_bytes):
    """parse XLSX bytes → list of {hostname, ip, raw}"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c) if c is not None else "" for c in rows[0]]
    h_idx, ip_idx = _find_columns(header)
    if h_idx is None and ip_idx is None:
        raise ValueError(f"找不到 hostname / ip 欄位 (header: {header})")
    out = []
    for r in rows[1:]:
        if not r or all(c is None or str(c).strip() == "" for c in r):
            continue
        hostname = str(r[h_idx]).strip() if h_idx is not None and len(r) > h_idx and r[h_idx] is not None else ""
        ip = str(r[ip_idx]).strip() if ip_idx is not None and len(r) > ip_idx and r[ip_idx] is not None else ""
        if hostname or ip:
            raw = {header[i]: (str(c) if c is not None else "") for i, c in enumerate(r) if i < len(header)}
            out.append({"hostname": hostname, "ip": ip, "raw": raw})
    return out


def parse_file(filename, file_bytes):
    fn = (filename or "").lower()
    if fn.endswith(".xlsx") or fn.endswith(".xlsm"):
        return parse_xlsx(file_bytes)
    return parse_csv(file_bytes)


def compare(parsed_rows):
    """跟 hosts collection 比對, 三欄分類

    Returns:
        {
            "excel_only": [{...}],     # 上傳有但 DB 沒
            "matched":    [{...}],      # 兩邊都有 (key=hostname or ip)
            "db_only":    [{...}],      # DB 有但上傳沒
            "stats": {excel_only_count, matched_count, db_only_count, excel_total, db_total}
        }
    """
    hosts = list(get_collection("hosts").find({}, {"_id": 0, "hostname": 1, "ip": 1, "ips": 1, "aliases": 1}))

    # 建 db 索引: hostname → host, alias → host, ip → host
    db_by_hostname = {h["hostname"].lower(): h for h in hosts if h.get("hostname")}
    db_by_alias = {a.lower(): h for h in hosts for a in (h.get("aliases") or [])}
    db_by_ip = {}
    for h in hosts:
        for ip in (h.get("ips") or []) + ([h["ip"]] if h.get("ip") else []):
            if ip:
                db_by_ip[ip] = h

    matched = []
    excel_only = []
    matched_db_hosts = set()  # 已配對到的 hostname

    for row in parsed_rows:
        hn = row.get("hostname", "").strip()
        ip = row.get("ip", "").strip()
        match_h = None

        if hn:
            match_h = db_by_hostname.get(hn.lower()) or db_by_alias.get(hn.lower())
        if not match_h and ip:
            match_h = db_by_ip.get(ip)

        if match_h:
            matched.append({
                "excel_hostname": hn,
                "excel_ip": ip,
                "db_hostname": match_h.get("hostname"),
                "db_ip": match_h.get("ip"),
                "match_by": ("hostname" if hn and (hn.lower() == match_h.get("hostname", "").lower() or hn.lower() in [a.lower() for a in (match_h.get("aliases") or [])]) else "ip"),
            })
            matched_db_hosts.add(match_h.get("hostname"))
        else:
            excel_only.append({"hostname": hn, "ip": ip, "raw": row.get("raw", {})})

    # DB only
    db_only = [
        {"hostname": h.get("hostname"), "ip": h.get("ip"), "ips": h.get("ips") or []}
        for h in hosts if h.get("hostname") not in matched_db_hosts
    ]

    return {
        "excel_only": excel_only,
        "matched": matched,
        "db_only": db_only,
        "stats": {
            "excel_total": len(parsed_rows),
            "db_total": len(hosts),
            "excel_only_count": len(excel_only),
            "matched_count": len(matched),
            "db_only_count": len(db_only),
        },
    }
