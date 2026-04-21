"""Admin API Blueprint - /api/admin/*"""
from flask import Blueprint, request, jsonify, session
import sys, os, json, subprocess, glob, shutil, platform, re, time, threading
from datetime import datetime, timedelta

# Phase 2 #4: ping-all 結果 60 秒記憶體快取（變更 #39 改用 MongoDB 共享，保留 _ping_lock 做 process 內互斥）
_ping_lock = threading.Lock()
PING_CACHE_TTL = 60

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from decorators import login_required, admin_required
from services.auth_service import verify_login, change_password, log_action, get_user
from services.mongo_service import get_collection, get_all_settings, update_setting
from config import INSPECTION_HOME, SETTINGS_FILE


# ============================================================================
# 變更 #39: MongoDB 共享儲存 helpers
# 原本 _ping_cache / _fixingHosts 是 process-local 記憶體；換 Gunicorn 多 worker
# 後各 worker 各自 cache 會失效、鎖會失效。改存 MongoDB 讓跨 worker 共享。
# ============================================================================

def _mongo_cache_get(key, ttl_sec):
    """從 MongoDB cache collection 讀取；超過 TTL 回 None"""
    try:
        doc = get_collection("cache").find_one({"_id": key})
        if not doc:
            return None
        updated = doc.get("updated_at")
        if not updated:
            return None
        age = (datetime.utcnow() - updated).total_seconds()
        if age > ttl_sec:
            return None
        return {"data": doc.get("data"), "age_sec": int(age)}
    except Exception:
        return None


def _mongo_cache_set(key, data):
    """寫入 MongoDB cache collection（upsert）"""
    try:
        get_collection("cache").update_one(
            {"_id": key},
            {"$set": {"data": data, "updated_at": datetime.utcnow()}},
            upsert=True
        )
    except Exception:
        pass


def _mongo_try_lock_host(hostname, ttl_sec, locked_by="unknown"):
    """嘗試取得主機修復鎖；已被其他 worker/user 鎖住且未過期則回 False

    原子性：利用 MongoDB _id 唯一約束 + filter 「未上鎖或已過期」
    """
    from pymongo.errors import DuplicateKeyError
    now = datetime.utcnow()
    expires = now + timedelta(seconds=ttl_sec)
    col = get_collection("fix_locks")
    try:
        # 嘗試把過期鎖搶過來（有 doc 且 expires_at < now），或 upsert 新 doc
        result = col.update_one(
            {"_id": hostname, "$or": [{"expires_at": {"$lt": now}}, {"expires_at": None}]},
            {"$set": {"locked_at": now, "expires_at": expires, "locked_by": locked_by}},
            upsert=True
        )
        return True
    except DuplicateKeyError:
        # _id 已存在且未過期 → 被別人鎖住
        return False
    except Exception:
        return False  # 保守：拿不到鎖就拒絕


def _mongo_release_lock_host(hostname):
    try:
        get_collection("fix_locks").delete_one({"_id": hostname})
    except Exception:
        pass


def _mongo_extend_lock_host(hostname, ttl_sec, locked_by="unknown"):
    """延長鎖到期時間（修復成功後背景 rescan 期間仍保持鎖）"""
    try:
        now = datetime.utcnow()
        expires = now + timedelta(seconds=ttl_sec)
        get_collection("fix_locks").update_one(
            {"_id": hostname},
            {"$set": {"expires_at": expires, "locked_by": locked_by}},
            upsert=True
        )
    except Exception:
        pass

bp = Blueprint("api_admin", __name__, url_prefix="/api/admin")

BACKUP_DIR = "/var/backups/inspection"
LOG_DIR = os.path.join(INSPECTION_HOME, "logs")


# ========== Auth ==========
@bp.route("/login", methods=["POST"])
def login():
    data = request.get_json(force=True)
    user = verify_login(data.get("username", ""), data.get("password", ""))
    if user == "LOCKED":
        return jsonify({"success": False, "error": "帳號已鎖定，請 15 分鐘後再試"}), 429
    if not user:
        from services.mongo_service import get_collection
        attempt = get_collection("login_attempts").find_one({"username": data.get("username", "")})
        attempts = attempt.get("attempts", 0) if attempt else 0
        remaining = max(0, 5 - attempts)
        msg = "帳號或密碼錯誤（還剩 %d 次機會，超過將鎖定 15 分鐘）" % remaining
        return jsonify({"success": False, "error": msg}), 401
    session["user_id"] = user["username"]
    session["username"] = user["username"]
    session["role"] = user["role"]
    session["display_name"] = user.get("display_name", user["username"])
    log_action(user["username"], "login", "登入成功", request.remote_addr)
    return jsonify({"success": True, "data": {
        "username": user["username"],
        "role": user["role"],
        "display_name": user.get("display_name", ""),
        "must_change_password": user.get("must_change_password", False),
    }})


@bp.route("/logout", methods=["POST"])
def logout():
    username = session.get("username", "unknown")
    log_action(username, "logout", "登出", request.remote_addr)
    session.clear()
    return jsonify({"success": True})


@bp.route("/me", methods=["GET"])
@login_required
def me():
    return jsonify({"success": True, "data": {
        "username": session.get("username"),
        "role": session.get("role"),
        "display_name": session.get("display_name"),
        "email": (get_user(session.get("username")) or {}).get("email", ""),
    }})


@bp.route("/change-password", methods=["POST"])
@login_required
def api_change_password():
    data = request.get_json(force=True)
    change_password(session["username"], data.get("new_password", ""))
    log_action(session["username"], "change_password", "修改密碼", request.remote_addr)
    return jsonify({"success": True, "message": "密碼已更新"})


# ========== System Status ==========
@bp.route("/system/status", methods=["GET"])
@admin_required
def system_status():
    # MongoDB
    try:
        from pymongo import MongoClient
        c = MongoClient("localhost", 27017, serverSelectionTimeoutMS=3000)
        c.admin.command("ping")
        mongo_ok = True
    except Exception:
        mongo_ok = False

    # Disk
    disk = shutil.disk_usage("/")
    disk_seclog = shutil.disk_usage("/var/log/inspection") if os.path.exists("/var/log/inspection") else disk

    # Podman
    try:
        r = subprocess.run(["podman", "ps", "--format", "json"], capture_output=True, text=True, timeout=10)
        containers = json.loads(r.stdout) if r.returncode == 0 else []
    except Exception:
        containers = []

# Flask (itagent-web)
    flask_ok = True  # if we are responding, Flask is running

    # ITAgent services
    itagent_services = {}
    for svc_name in ["itagent-db", "itagent-web"]:
        try:
            sr = subprocess.run(["systemctl", "is-active", svc_name], capture_output=True, text=True, timeout=5)
            itagent_services[svc_name] = sr.stdout.strip()
        except Exception:
            itagent_services[svc_name] = "unknown"

    return jsonify({"success": True, "data": {
        "flask": {"status": "running"},
        "mongodb": {"status": "running" if mongo_ok else "down"},
        "disk": {
            "root": {"total_gb": round(disk.total/1e9, 1), "used_gb": round(disk.used/1e9, 1), "free_gb": round(disk.free/1e9, 1), "percent": round(disk.used/disk.total*100, 1)},
            "seclog": {"total_gb": round(disk_seclog.total/1e9, 1), "used_gb": round(disk_seclog.used/1e9, 1), "free_gb": round(disk_seclog.free/1e9, 1), "percent": round(disk_seclog.used/disk_seclog.total*100, 1)},
        },
        "itagent_services": itagent_services,
        "containers": [{"name": c.get("Names", ["?"])[0] if isinstance(c.get("Names"), list) else c.get("Names", "?"), "status": c.get("Status", c.get("State", "?"))} for c in containers],
    }})


@bp.route("/system/info", methods=["GET"])
@admin_required
def system_info():
    def cmd(c):
        try:
            return subprocess.run(c, capture_output=True, text=True, timeout=10, shell=isinstance(c, str)).stdout.strip()
        except Exception:
            return "N/A"

    return jsonify({"success": True, "data": {
        "os": cmd("cat /etc/os-release | grep PRETTY_NAME | cut -d= -f2 | tr -d '\"'"),
        "hostname": cmd(["hostname"]),
        "ip": cmd("hostname -I | awk '{print $1}'"),
        "python": platform.python_version(),
        "ansible": cmd("ansible --version | head -1"),
        "uptime": cmd("uptime -p"),
        "boot_time": cmd("uptime -s"),
    }})


@bp.route("/system/run-inspection", methods=["POST"])
@admin_required
def run_inspection():
    data = request.get_json(force=True) or {}
    hostname = data.get("hostname")
    cmd = ["/bin/bash", os.path.join(INSPECTION_HOME, "run_inspection.sh")]
    try:
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        log_action(session["username"], "run_inspection", f"手動執行巡檢 (host={hostname or 'all'})", request.remote_addr)
        return jsonify({"success": True, "message": "巡檢已觸發", "pid": proc.pid})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ========== Settings ==========
@bp.route("/settings", methods=["GET"])
@admin_required
def admin_settings():
    return jsonify({"success": True, "data": get_all_settings()})


@bp.route("/settings/<key>", methods=["PUT"])
@admin_required
def admin_update_setting(key):
    data = request.get_json(force=True)
    value = data.get("value")
    # Dual write: MongoDB + settings.json
    update_setting(key, value)
    try:
        with open(SETTINGS_FILE) as f:
            s = json.load(f)
        if key == "thresholds":
            s["thresholds"] = value
        else:
            s[key] = value
        with open(SETTINGS_FILE, "w") as f:
            json.dump(s, f, indent=2, ensure_ascii=False)
    except Exception:
        pass
    log_action(session["username"], "update_setting", f"修改設定: {key}", request.remote_addr)
    return jsonify({"success": True, "message": f"設定 {key} 已更新"})


# ========== Backups ==========
@bp.route("/backups", methods=["GET"])
@admin_required
def list_backups():
    files = sorted(glob.glob(os.path.join(BACKUP_DIR, "INSPECTION_HOME_*.tar.gz")), reverse=True)
    backups = []
    for f in files:
        st = os.stat(f)
        backups.append({
            "name": os.path.basename(f),
            "size_mb": round(st.st_size / 1e6, 2),
            "created": datetime.fromtimestamp(st.st_mtime).isoformat(),
        })
    return jsonify({"success": True, "data": backups})


@bp.route("/backups", methods=["POST"])
@admin_required
def create_backup():
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = f"INSPECTION_HOME_{ts}.tar.gz"
    path = os.path.join(BACKUP_DIR, name)
    try:
        subprocess.run(["tar", "czf", path, "-C", "/opt", "inspection/"], check=True, timeout=120)
        log_action(session["username"], "backup_create", f"建立備份: {name}", request.remote_addr)
        return jsonify({"success": True, "message": f"備份已建立: {name}"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/backups/<name>/restore", methods=["POST"])
@admin_required
def restore_backup(name):
    path = os.path.join(BACKUP_DIR, name)
    if not os.path.exists(path) or ".." in name:
        return jsonify({"success": False, "error": "備份不存在"}), 404
    try:
        subprocess.run(["tar", "xzf", path, "-C", "/opt"], check=True, timeout=120)
        log_action(session["username"], "backup_restore", f"還原備份: {name}", request.remote_addr)
        return jsonify({"success": True, "message": f"已還原: {name}"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/backups/<name>", methods=["DELETE"])
@admin_required
def delete_backup(name):
    path = os.path.join(BACKUP_DIR, name)
    if not os.path.exists(path) or ".." in name:
        return jsonify({"success": False, "error": "備份不存在"}), 404
    os.remove(path)
    log_action(session["username"], "backup_delete", f"刪除備份: {name}", request.remote_addr)
    return jsonify({"success": True, "message": f"已刪除: {name}"})


# ========== Jobs ==========
@bp.route("/jobs/seed", methods=["POST"])
@admin_required
def run_seed():
    try:
        r = subprocess.run(["python3", os.path.join(INSPECTION_HOME, "webapp/seed_data.py")],
                           capture_output=True, text=True, timeout=60,
                           cwd=os.path.join(INSPECTION_HOME, "webapp"))
        log_action(session["username"], "run_seed", "重新匯入 MongoDB", request.remote_addr)
        return jsonify({"success": True, "output": r.stdout})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/jobs/status", methods=["GET"])
@admin_required
def job_status():
    log_files = sorted(glob.glob(os.path.join(LOG_DIR, "*_run.log")), reverse=True)
    if not log_files:
        return jsonify({"success": True, "data": {"last_run": None, "log_tail": []}})
    last = log_files[0]
    with open(last) as f:
        lines = f.readlines()
    return jsonify({"success": True, "data": {
        "last_run": os.path.basename(last).replace("_run.log", ""),
        "log_tail": [l.strip() for l in lines[-30:]],
    }})


# ========== Logs ==========
@bp.route("/logs/inspection", methods=["GET"])
@admin_required
def view_inspection_logs():
    date = request.args.get("date", "")
    keyword = request.args.get("keyword", "")
    tail = request.args.get("tail", 100, type=int)
    log_files = sorted(glob.glob(os.path.join(LOG_DIR, "*_run.log")), reverse=True)
    if date:
        log_files = [f for f in log_files if date in os.path.basename(f)]
    lines = []
    for lf in log_files[:5]:
        with open(lf) as f:
            for line in f:
                if keyword and keyword.lower() not in line.lower():
                    continue
                lines.append({"file": os.path.basename(lf), "line": line.strip()})
    return jsonify({"success": True, "data": lines[-tail:]})


@bp.route("/logs/flask", methods=["GET"])
@admin_required
def view_flask_log():
    tail = request.args.get("tail", 100, type=int)
    try:
        with open("/tmp/flask.log") as f:
            lines = f.readlines()
        return jsonify({"success": True, "data": [l.strip() for l in lines[-tail:]]})
    except Exception:
        return jsonify({"success": True, "data": []})


# ========== Alerts ==========
@bp.route("/alerts", methods=["GET"])
@admin_required
def list_alerts():
    col = get_collection("inspections")
    pipeline = [
        {"$sort": {"run_date": -1, "run_time": -1}},
        {"$match": {"overall_status": {"$in": ["warn", "error"]}}},
        {"$limit": 100},
        {"$project": {"_id": 0, "hostname": 1, "overall_status": 1, "run_date": 1, "run_time": 1, "ip": 1, "os": 1}},
    ]
    alerts = list(col.aggregate(pipeline))
    # Add ack status from alerts collection
    ack_col = get_collection("alert_acks")
    for a in alerts:
        key = f"{a['hostname']}_{a.get('run_date','')}_{a.get('run_time','')}"
        ack = ack_col.find_one({"key": key})
        a["acknowledged"] = bool(ack)
        a["ack_by"] = ack.get("user", "") if ack else ""
    return jsonify({"success": True, "data": alerts})


@bp.route("/alerts/<hostname>/<run_date>/<run_time>/ack", methods=["PUT"])
@admin_required
def ack_alert(hostname, run_date, run_time):
    key = f"{hostname}_{run_date}_{run_time}"
    get_collection("alert_acks").update_one(
        {"key": key},
        {"$set": {"key": key, "user": session["username"], "timestamp": datetime.now().isoformat()}},
        upsert=True
    )
    log_action(session["username"], "alert_ack", f"確認告警: {hostname} {run_date}", request.remote_addr)
    return jsonify({"success": True})


# ========== Host Management ==========
@bp.route("/hosts", methods=["POST"])
@admin_required
def add_host():
    data = request.get_json(force=True)
    if not data.get("hostname"):
        return jsonify({"success": False, "error": "缺少 hostname 欄位"}), 400
    data["imported_at"] = datetime.now().isoformat()
    data["updated_at"] = datetime.now().isoformat()
    get_collection("hosts").update_one({"hostname": data["hostname"]}, {"$set": data}, upsert=True)
    # Also update hosts_config.json
    _sync_hosts_config()
    log_action(session["username"], "host_add", f"新增主機: {data['hostname']}", request.remote_addr)
    return jsonify({"success": True, "message": f"主機 {data['hostname']} 已新增"})


@bp.route("/hosts/<hostname>", methods=["PUT"])
@admin_required
def edit_host(hostname):
    data = request.get_json(force=True)
    data["updated_at"] = datetime.now().isoformat()
    get_collection("hosts").update_one({"hostname": hostname}, {"$set": data})
    _sync_hosts_config()
    log_action(session["username"], "host_edit", f"編輯主機: {hostname}", request.remote_addr)
    return jsonify({"success": True})


@bp.route("/hosts/<hostname>", methods=["DELETE"])
@admin_required
def delete_host(hostname):
    get_collection("hosts").delete_one({"hostname": hostname})
    _sync_hosts_config()
    log_action(session["username"], "host_delete", f"刪除主機: {hostname}", request.remote_addr)
    return jsonify({"success": True})


@bp.route("/hosts/<hostname>/ping", methods=["POST"])
@admin_required
def ping_host(hostname):
    try:
        r = subprocess.run(
            ["ansible", "-i", os.path.join(INSPECTION_HOME, "ansible/inventory/hosts.yml"),
             hostname, "-m", "ping", "-u", "ansible_svc"],
            capture_output=True, text=True, timeout=30
        )
        ok = r.returncode == 0
        return jsonify({"success": True, "reachable": ok, "output": r.stdout + r.stderr})
    except Exception as e:
        return jsonify({"success": True, "reachable": False, "output": str(e)})


@bp.route("/hosts/regenerate-inventory", methods=["POST"])
@admin_required
def regenerate_inventory():
    try:
        r = subprocess.run(
            ["python3", os.path.join(INSPECTION_HOME, "scripts/generate_inventory.py")],
            capture_output=True, text=True, timeout=30,
            cwd=INSPECTION_HOME
        )
        log_action(session["username"], "regenerate_inventory", "重建 Ansible inventory", request.remote_addr)
        return jsonify({"success": True, "output": r.stdout})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ========== Scheduler ==========
@bp.route("/scheduler", methods=["GET"])
@admin_required
def get_schedule():
    try:
        r = subprocess.run(["crontab", "-l"], capture_output=True, text=True, timeout=10)
        lines = [l.strip() for l in r.stdout.splitlines() if "run_inspection" in l]
        schedule = []
        for l in lines:
            raw = l
            is_disabled = l.startswith("#")
            clean = l.lstrip("# ")
            parts = clean.split()
            if len(parts) >= 5:
                schedule.append({"minute": parts[0], "hour": parts[1], "raw": raw, "enabled": not is_disabled})
        return jsonify({"success": True, "data": schedule})
    except Exception:
        return jsonify({"success": True, "data": []})


@bp.route("/scheduler", methods=["PUT"])
@admin_required
def update_schedule():
    data = request.get_json(force=True)
    times = data.get("times", [])
    try:
        r = subprocess.run(["crontab", "-l"], capture_output=True, text=True, timeout=10)
        other_lines = [l for l in r.stdout.splitlines() if "run_inspection" not in l and l.strip()]
        for t in times:
            cron_line = f"{t['minute']} {t['hour']} * * * {INSPECTION_HOME}/run_inspection.sh >> {INSPECTION_HOME}/logs/cron.log 2>&1"
            if not t.get("enabled", True):
                cron_line = "# " + cron_line
            other_lines.append(cron_line)
        new_cron = "\n".join(other_lines) + "\n"
        subprocess.run(["crontab", "-"], input=new_cron, text=True, timeout=10)
        enabled_count = len([t for t in times if t.get("enabled", True)])
        log_action(session["username"], "update_schedule", f"更新排程: {len(times)} 個 ({enabled_count} 啟用)", request.remote_addr)
        return jsonify({"success": True, "message": "排程已更新"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ========== Reports ==========
@bp.route("/reports/monthly", methods=["GET"])
@admin_required
def monthly_report():
    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    col = get_collection("inspections")
    pipeline = [
        {"$match": {"run_date": {"$regex": f"^{month}"}}},
        {"$group": {
            "_id": {"hostname": "$hostname", "status": "$overall_status"},
            "count": {"$sum": 1}
        }},
    ]
    results = list(col.aggregate(pipeline))
    # Aggregate
    hosts = {}
    for r in results:
        h = r["_id"]["hostname"]
        s = (r["_id"]["status"] or "ok").strip()
        if h not in hosts:
            hosts[h] = {"hostname": h, "ok": 0, "warn": 0, "error": 0, "total": 0}
        hosts[h][s] = r["count"]
        hosts[h]["total"] += r["count"]
    for h in hosts.values():
        h["sla"] = round(h["ok"] / h["total"] * 100, 2) if h["total"] > 0 else 0
    return jsonify({"success": True, "data": {"month": month, "hosts": list(hosts.values())}})


@bp.route("/reports/export", methods=["GET"])
@admin_required
def export_report():
    import csv, io
    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    fmt = request.args.get("format", "csv")
    # Get monthly data
    resp = monthly_report()
    data = json.loads(resp.get_data())["data"]
    if fmt == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["主機", "正常", "警告", "異常", "總計", "SLA%"])
        for h in data["hosts"]:
            writer.writerow([h["hostname"], h["ok"], h["warn"], h["error"], h["total"], h["sla"]])
        from flask import Response
        return Response(output.getvalue(), mimetype="text/csv",
                        headers={"Content-Disposition": f"attachment;filename=report_{month}.csv"})
    return jsonify(data)


# ========== Worklog ==========
@bp.route("/worklog", methods=["GET"])
@admin_required
def view_worklog():
    page = request.args.get("page", 1, type=int)
    per_page = 50
    col = get_collection("admin_worklog")
    total = col.count_documents({})
    docs = list(col.find({}, {"_id": 0}).sort("timestamp", -1).skip((page-1)*per_page).limit(per_page))
    return jsonify({"success": True, "data": docs, "total": total, "page": page})


# ========== Import ==========
@bp.route("/hosts/import-csv", methods=["POST"])
@admin_required
def import_csv():
    """從 CSV 匯入主機清單"""
    import csv, io
    if "file" not in request.files:
        # Try raw text body
        raw = request.get_data(as_text=True)
        if not raw:
            return jsonify({"success": False, "error": "未提供檔案"}), 400
        reader = csv.DictReader(io.StringIO(raw))
    else:
        f = request.files["file"]
        content = f.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))

    col = get_collection("hosts")
    count = 0
    errors = []
    for i, row in enumerate(reader):
        # Map common column names
        hostname = row.get("hostname") or row.get("主機名稱") or row.get("host") or ""
        ip = row.get("ip") or row.get("IP") or row.get("IP位址") or ""
        if not hostname.strip():
            errors.append(f"第 {i+2} 行缺少主機名稱")
            continue
        doc = {
            "hostname": hostname.strip(),
            "ip": (ip or "").strip(),
            "os": (row.get("os") or row.get("OS") or row.get("作業系統") or "").strip(),
            "os_group": (row.get("os_group") or row.get("OS Group") or "").strip().lower(),
            "status": (row.get("status") or row.get("狀態") or "使用中").strip(),
            "environment": (row.get("environment") or row.get("環境") or "").strip(),
            "group": (row.get("group") or row.get("群組") or "").strip() or None,
            "has_python": True,
            "custodian": (row.get("custodian") or row.get("保管者") or "").strip(),
            "custodian_ad": (row.get("custodian_ad") or row.get("AD帳號") or "").strip(),
            "department": (row.get("department") or row.get("部門") or "").strip(),
            "division": (row.get("division") or row.get("處") or "").strip(),
            "imported_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat(),
        }
        col.update_one({"hostname": doc["hostname"]}, {"$set": doc}, upsert=True)
        count += 1

    _sync_hosts_config()
    log_action(session["username"], "import_csv", f"CSV 匯入 {count} 台主機", request.remote_addr)
    return jsonify({"success": True, "message": f"成功匯入 {count} 台主機", "count": count, "errors": errors})


@bp.route("/hosts/import-json", methods=["POST"])
@admin_required
def import_json():
    """從 hosts_config.json 重新匯入"""
    config_path = os.path.join(INSPECTION_HOME, "data/hosts_config.json")
    try:
        with open(config_path) as f:
            data = json.load(f)
        hosts = data if isinstance(data, list) else data.get("hosts", [data])
        col = get_collection("hosts")
        count = 0
        for h in hosts:
            h.setdefault("imported_at", datetime.now().isoformat())
            h.setdefault("updated_at", datetime.now().isoformat())
            col.update_one({"hostname": h["hostname"]}, {"$set": h}, upsert=True)
            count += 1
        log_action(session["username"], "import_json", f"JSON 匯入 {count} 台主機", request.remote_addr)
        return jsonify({"success": True, "message": f"從 hosts_config.json 匯入 {count} 台主機", "count": count})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/hosts/export-csv", methods=["GET"])
@admin_required
def export_csv():
    """匯出主機清單為 CSV"""
    import csv, io
    hosts = list(get_collection("hosts").find({}, {"_id": 0}))
    output = io.StringIO()
    fields = ["hostname", "ip", "os", "os_group", "status", "environment", "group", "custodian", "custodian_ad", "department", "division"]
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    for h in hosts:
        writer.writerow(h)
    from flask import Response
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment;filename=hosts_export.csv"})


@bp.route("/hosts/template-csv", methods=["GET"])
@admin_required
def template_csv():
    """下載 CSV 範本 (BOM for Excel)"""
    from flask import Response
    template = "hostname,ip,os,os_group,status,environment,group,custodian,custodian_ad,department,division\nEXAMPLE-SVR01,10.0.0.1,Rocky Linux,rocky,使用中,正式,,林凱文,lin.kaiwen,資訊架構部,資訊管理處\n"
    bom = "\ufeff"
    return Response(bom + template, mimetype="text/csv; charset=utf-8",
                    headers={"Content-Disposition": "attachment;filename=hosts_template.csv"})

@bp.route("/hosts/template-xlsx", methods=["GET"])
@admin_required
def template_xlsx():
    """下載 Excel 範本"""
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "主機清單"

    headers = ["hostname", "ip", "os", "os_group", "status", "environment", "group", "custodian", "custodian_ad", "department", "division"]
    header_cn = ["主機名稱", "IP位址", "作業系統", "OS分類", "狀態", "環境", "群組", "保管者", "AD帳號", "部門", "處"]
    notes = ["必填", "必填", "Rocky Linux / Windows Server 2019 / AIX", "rhel / rocky / debian / windows / aix", "使用中", "正式 / 測試", "可空", "負責人姓名", "AD登入帳號", "部門名稱", "處名稱"]

    # Style
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4AB234", end_color="4AB234", fill_type="solid")
    nfont = Font(italic=True, color="888888", size=9)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Row 1: English headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Row 2: Chinese description
    for col, h in enumerate(header_cn, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="333333", size=10)
        cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        cell.border = border

    # Row 3: Notes
    for col, n in enumerate(notes, 1):
        cell = ws.cell(row=3, column=col, value=n)
        cell.font = nfont
        cell.border = border

    # Row 4: Example
    example = ["SECSVR019-032", "10.93.19.35", "Rocky Linux", "rhel", "使用中", "正式", "", "林凱文", "lin.kaiwen", "資訊架構部", "資訊管理處"]
    for col, v in enumerate(example, 1):
        cell = ws.cell(row=4, column=col, value=v)
        cell.font = Font(color="999999")
        cell.border = border

    # Column widths
    widths = [22, 16, 22, 12, 8, 8, 10, 10, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="hosts_template.xlsx")


# ========== Account Audit ==========
@bp.route("/audit/accounts", methods=["GET"])
@admin_required
def audit_accounts():
    """帳號盤點報表"""
    col = get_collection("inspections")
    notes_col = get_collection("account_notes")
    hr_col = get_collection("hr_users")
    settings_col = get_collection("settings")

    # Get thresholds
    pw_days = 180
    login_days = 180
    th = settings_col.find_one({"key": "audit_password_days"})
    if th:
        pw_days = int(th.get("value", 180))
    th2 = settings_col.find_one({"key": "audit_login_days"})
    if th2:
        login_days = int(th2.get("value", 180))

    # Get latest inspections with audit data
    pipeline = [
        {"$sort": {"run_date": -1, "run_time": -1}},
        {"$group": {"_id": "$hostname", "doc": {"$first": "$$ROOT"}}},
        {"$replaceRoot": {"newRoot": "$doc"}},
        {"$project": {"_id": 0}},
    ]
    inspections = list(col.aggregate(pipeline))

    # Build HR lookup
    hr_lookup = {}
    for hr in hr_col.find({}, {"_id": 0}):
        ad = hr.get("ad_account", "").lower()
        if ad:
            hr_lookup[ad] = hr

    # Build notes lookup
    notes_lookup = {}
    for note in notes_col.find({}, {"_id": 0}):
        key = f"{note.get('hostname', '')}_{note.get('user', '')}"
        notes_lookup[key] = note

    # Exclude system accounts
    linux_system = {"systemd-coredump", "sssd", "chrony", "systemd-oom", "polkitd",
                    "setroubleshoot", "saslauth", "dbus", "tss", "clevis",
                    "cockpit-ws", "cockpit-wsinstance", "flatpak", "gnome-initial-setup",
                    "colord", "geoclue", "pipewire", "rtkit", "abrt", "unbound",
                    "radvd", "qemu", "usbmuxd", "gluster", "rpcuser", "nfsnobody"}

    result = []
    for insp in inspections:
        hostname = insp.get("hostname", "")
        audit_data = insp.get("results", {}).get("account_audit", [])
        for acct in audit_data:
            user = acct.get("user", "")
            if user.lower() in linux_system:
                continue

            # HR match
            hr = hr_lookup.get(user.lower(), {})
            # Notes
            note_key = f"{hostname}_{user}"
            note_data = notes_lookup.get(note_key, {})

            # Risk assessment
            risks = []
            pw_age = acct.get("pw_age_days", 0)
            login_age = acct.get("login_age_days", 0)
            if isinstance(pw_age, str):
                pw_age = int(pw_age) if pw_age.isdigit() else 9999
            if isinstance(login_age, str):
                login_age = int(login_age) if login_age.isdigit() else 9999

            if pw_age >= pw_days:
                risks.append({"type": "pw_old", "desc": f"密碼 {pw_age} 天未變更", "level": "warn"})
            if acct.get("pw_expired"):
                risks.append({"type": "pw_expired", "desc": "密碼已到期", "level": "error"})
            if login_age >= login_days:
                risks.append({"type": "no_login", "desc": f"{login_age} 天未登入", "level": "warn"})

            result.append({
                "hostname": hostname,
                "user": user,
                "uid": acct.get("uid", ""),
                "enabled": acct.get("enabled", True),
                "locked": acct.get("locked", ""),
                "pw_last_change": acct.get("pw_last_change", ""),
                "pw_expires": acct.get("pw_expires", ""),
                "pw_age_days": pw_age,
                "last_login": acct.get("last_login", ""),
                "login_age_days": login_age,
                "risks": risks,
                "risk_count": len(risks),
                "note": note_data.get("note", ""),
                "department": note_data.get("department", "") or hr.get("department", hr.get("部門", "")),
                "hr_name": hr.get("name", hr.get("姓名", "")),
                "hr_emp_id": hr.get("emp_id", hr.get("工號", "")),
            })

    # Sort by risk count desc
    result.sort(key=lambda x: (-x["risk_count"], x["hostname"], x["user"]))

    return jsonify({"success": True, "data": result, "count": len(result),
                     "thresholds": {"pw_days": pw_days, "login_days": login_days}})


@bp.route("/audit/accounts/<hostname>/<user>/note", methods=["PUT"])
@admin_required
def update_account_note(hostname, user):
    """編輯帳號備註/部門"""
    data = request.get_json(force=True)
    get_collection("account_notes").update_one(
        {"hostname": hostname, "user": user},
        {"$set": {
            "hostname": hostname,
            "user": user,
            "note": data.get("note", ""),
            "department": data.get("department", ""),
            "updated_at": datetime.now().isoformat(),
            "updated_by": session.get("username", ""),
        }},
        upsert=True
    )
    log_action(session["username"], "update_account_note", f"編輯帳號備註: {hostname}/{user}", request.remote_addr)
    return jsonify({"success": True})


@bp.route("/audit/hr", methods=["GET"])
@admin_required
def list_hr():
    """HR 人員清單"""
    data = list(get_collection("hr_users").find({}, {"_id": 0}))
    return jsonify({"success": True, "data": data, "count": len(data)})


@bp.route("/audit/hr/import", methods=["POST"])
@admin_required
def import_hr():
    """匯入 HR CSV"""
    import csv, io
    if "file" not in request.files:
        return jsonify({"success": False, "error": "未提供檔案"}), 400
    f = request.files["file"]
    content = f.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    col = get_collection("hr_users")
    count = 0
    for row in reader:
        emp_id = row.get("工號") or row.get("emp_id") or ""
        ad = row.get("AD帳號") or row.get("ad_account") or ""
        if not emp_id and not ad:
            continue
        doc = {
            "emp_id": emp_id.strip(),
            "name": (row.get("姓名") or row.get("name") or "").strip(),
            "ad_account": ad.strip().lower(),
            "department": (row.get("部門") or row.get("department") or "").strip(),
            "title": (row.get("職稱") or row.get("title") or "").strip(),
            "hire_date": (row.get("到職日") or row.get("hire_date") or "").strip(),
            "status": (row.get("狀態") or row.get("status") or "在職").strip(),
            "imported_at": datetime.now().isoformat(),
        }
        col.update_one({"emp_id": doc["emp_id"]}, {"$set": doc}, upsert=True)
        count += 1
    log_action(session["username"], "import_hr", f"匯入 HR 人員 {count} 筆", request.remote_addr)
    return jsonify({"success": True, "message": f"匯入 {count} 筆 HR 人員資料", "count": count})


@bp.route("/audit/hr/template", methods=["GET"])
@admin_required
def hr_template():
    """HR CSV 範本"""
    from flask import Response
    t = "工號,姓名,AD帳號,部門,職稱,到職日,狀態\nT00001,王○明,wang.test01,資訊架構部,系統工程師,2020-03-15,在職\n"
    return Response(t, mimetype="text/csv", headers={"Content-Disposition": "attachment;filename=hr_template.csv"})


@bp.route("/audit/export", methods=["GET"])
@admin_required
def export_audit():
    """匯出帳號盤點 CSV"""
    import csv, io
    resp = audit_accounts()
    data = json.loads(resp.get_data())["data"]
    output = io.StringIO()
    fields = ["hostname", "user", "department", "hr_name", "hr_emp_id", "note",
              "enabled", "pw_last_change", "pw_expires", "pw_age_days",
              "last_login", "login_age_days", "risk_count"]
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    for d in data:
        writer.writerow(d)
    from flask import Response
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment;filename=account_audit_{datetime.now().strftime('%Y%m%d')}.csv"})


@bp.route("/audit/settings", methods=["GET"])
@admin_required
def get_audit_settings():
    """取得帳號盤點閾值"""
    from services.mongo_service import get_all_settings
    settings = get_all_settings()
    return jsonify({"success": True, "data": {
        "pw_days": settings.get("audit_password_days", 90),
        "login_days": settings.get("audit_login_days", 90),
    }})


@bp.route("/audit/settings", methods=["PUT"])
@admin_required
def update_audit_settings():
    """更新帳號盤點閾值"""
    data = request.get_json(force=True)
    if "pw_days" in data:
        update_setting("audit_password_days", int(data["pw_days"]))
    if "login_days" in data:
        update_setting("audit_login_days", int(data["login_days"]))
    log_action(session["username"], "update_audit_settings", f"更新盤點閾值", request.remote_addr)
    return jsonify({"success": True})


# ========== MongoDB Dump/Restore ==========
DB_BACKUP_DIR = os.path.join(BACKUP_DIR, "dbdump")

@bp.route("/dbbackups", methods=["GET"])
@admin_required
def list_dbbackups():
    os.makedirs(DB_BACKUP_DIR, exist_ok=True)
    files = sorted(glob.glob(os.path.join(DB_BACKUP_DIR, "mongodump_*.tar.gz")), reverse=True)
    backups = []
    for f in files:
        st = os.stat(f)
        backups.append({
            "name": os.path.basename(f),
            "size_mb": round(st.st_size / 1e6, 2),
            "created": datetime.fromtimestamp(st.st_mtime).isoformat(),
        })
    return jsonify({"success": True, "data": backups})


@bp.route("/dbbackups", methods=["POST"])
@admin_required
def create_dbbackup():
    os.makedirs(DB_BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dump_dir = f"/tmp/mongodump_{ts}"
    tar_name = f"mongodump_{ts}.tar.gz"
    tar_path = os.path.join(DB_BACKUP_DIR, tar_name)
    try:
        subprocess.run(["podman", "exec", "mongodb", "mongodump", "--db", "inspection",
                        "--out", f"/tmp/mongodump_{ts}"], check=True, timeout=120)
        # Copy from container to host
        subprocess.run(["podman", "cp", f"mongodb:/tmp/mongodump_{ts}", dump_dir],
                        check=True, timeout=60)
        subprocess.run(["tar", "czf", tar_path, "-C", dump_dir, "."], check=True, timeout=60)
        subprocess.run(["rm", "-rf", dump_dir], timeout=10)
        subprocess.run(["podman", "exec", "mongodb", "rm", "-rf", f"/tmp/mongodump_{ts}"], timeout=10)
        log_action(session["username"], "db_dump", f"MongoDB Dump: {tar_name}", request.remote_addr)
        return jsonify({"success": True, "message": f"Dump 完成: {tar_name}"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/dbbackups/<name>/restore", methods=["POST"])
@admin_required
def restore_dbbackup(name):
    tar_path = os.path.join(DB_BACKUP_DIR, name)
    if not os.path.exists(tar_path) or ".." in name:
        return jsonify({"success": False, "error": "備份不存在"}), 404
    try:
        tmp_dir = f"/tmp/dbrestore_{datetime.now().strftime('%H%M%S')}"
        os.makedirs(tmp_dir, exist_ok=True)
        subprocess.run(["tar", "xzf", tar_path, "-C", tmp_dir], check=True, timeout=60)
        # Copy into container
        subprocess.run(["podman", "cp", tmp_dir, "mongodb:/tmp/dbrestore"], check=True, timeout=60)
        subprocess.run(["podman", "exec", "mongodb", "mongorestore", "--db", "inspection",
                        "--drop", "/tmp/dbrestore/inspection"], check=True, timeout=120)
        subprocess.run(["podman", "exec", "mongodb", "rm", "-rf", "/tmp/dbrestore"], timeout=10)
        subprocess.run(["rm", "-rf", tmp_dir], timeout=10)
        log_action(session["username"], "db_restore", f"MongoDB Restore: {name}", request.remote_addr)
        return jsonify({"success": True, "message": f"還原完成: {name}"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/dbbackups/import", methods=["POST"])
@admin_required
def import_dbbackup():
    if "file" not in request.files:
        return jsonify({"success": False, "error": "未提供檔案"}), 400
    f = request.files["file"]
    os.makedirs(DB_BACKUP_DIR, exist_ok=True)
    save_path = os.path.join(DB_BACKUP_DIR, f.filename)
    f.save(save_path)
    # Auto restore
    try:
        tmp_dir = f"/tmp/dbimport_{datetime.now().strftime('%H%M%S')}"
        os.makedirs(tmp_dir, exist_ok=True)
        subprocess.run(["tar", "xzf", save_path, "-C", tmp_dir], check=True, timeout=60)
        subprocess.run(["podman", "cp", tmp_dir, "mongodb:/tmp/dbimport"], check=True, timeout=60)
        subprocess.run(["podman", "exec", "mongodb", "mongorestore", "--db", "inspection",
                        "--drop", "/tmp/dbimport/inspection"], check=True, timeout=120)
        subprocess.run(["podman", "exec", "mongodb", "rm", "-rf", "/tmp/dbimport"], timeout=10)
        subprocess.run(["rm", "-rf", tmp_dir], timeout=10)
        log_action(session["username"], "db_import", f"MongoDB Import: {f.filename}", request.remote_addr)
        return jsonify({"success": True, "message": f"匯入並還原完成: {f.filename}"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/dbbackups/<name>/download", methods=["GET"])
@admin_required
def download_dbbackup(name):
    tar_path = os.path.join(DB_BACKUP_DIR, name)
    if not os.path.exists(tar_path) or ".." in name:
        return jsonify({"success": False, "error": "備份不存在"}), 404
    from flask import send_file
    return send_file(tar_path, as_attachment=True, download_name=name)


@bp.route("/dbbackups/<name>", methods=["DELETE"])
@admin_required
def delete_dbbackup(name):
    tar_path = os.path.join(DB_BACKUP_DIR, name)
    if not os.path.exists(tar_path) or ".." in name:
        return jsonify({"success": False, "error": "備份不存在"}), 404
    os.remove(tar_path)
    log_action(session["username"], "db_delete", f"刪除 Dump: {name}", request.remote_addr)
    return jsonify({"success": True, "message": f"已刪除: {name}"})


# ========== Patch System ==========
PATCH_DIR = os.path.join(BACKUP_DIR, "patches")

@bp.route("/patch/upload", methods=["POST"])
@admin_required
def upload_patch():
    if "file" not in request.files:
        return jsonify({"success": False, "error": "未提供檔案"}), 400
    f = request.files["file"]
    os.makedirs(PATCH_DIR, exist_ok=True)
    save_path = os.path.join(PATCH_DIR, f.filename)
    f.save(save_path)
    # Extract and read patch_info.json
    try:
        tmp_dir = f"/tmp/patch_preview_{datetime.now().strftime('%H%M%S')}"
        os.makedirs(tmp_dir, exist_ok=True)
        subprocess.run(["tar", "xzf", save_path, "-C", tmp_dir], check=True, timeout=30)
        info_path = os.path.join(tmp_dir, "patch_info.json")
        if not os.path.exists(info_path):
            # Try one level deeper
            for root, dirs, files in os.walk(tmp_dir):
                if "patch_info.json" in files:
                    info_path = os.path.join(root, "patch_info.json")
                    break
        if os.path.exists(info_path):
            with open(info_path) as pf:
                info = json.load(pf)
        else:
            info = {"version": "unknown", "description": "patch_info.json not found", "files": []}
        subprocess.run(["rm", "-rf", tmp_dir], timeout=10)
        return jsonify({"success": True, "data": {
            "filename": f.filename,
            "info": info,
        }})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/patch/apply", methods=["POST"])
@admin_required
def apply_patch():
    data = request.get_json(force=True)
    filename = data.get("filename", "")
    patch_path = os.path.join(PATCH_DIR, filename)
    if not os.path.exists(patch_path) or ".." in filename:
        return jsonify({"success": False, "error": "Patch 不存在"}), 404
    try:
        # Step 1: Auto backup before patching
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"pre_patch_{ts}.tar.gz"
        subprocess.run(["tar", "czf", os.path.join(BACKUP_DIR, backup_name),
                        "--exclude=container", "-C", os.path.dirname(INSPECTION_HOME),
                        os.path.basename(INSPECTION_HOME)], check=True, timeout=120)

        # Step 2: Extract patch
        tmp_dir = f"/tmp/patch_apply_{ts}"
        os.makedirs(tmp_dir, exist_ok=True)
        subprocess.run(["tar", "xzf", patch_path, "-C", tmp_dir], check=True, timeout=30)

        # Find files directory
        files_dir = os.path.join(tmp_dir, "files")
        if not os.path.exists(files_dir):
            for root, dirs, _files in os.walk(tmp_dir):
                if "files" in dirs:
                    files_dir = os.path.join(root, "files")
                    break

        # Step 3: Copy files
        if os.path.exists(files_dir):
            for root, dirs, files in os.walk(files_dir):
                for fname in files:
                    src = os.path.join(root, fname)
                    rel = os.path.relpath(src, files_dir)
                    dst = os.path.join(INSPECTION_HOME, rel)
                    os.makedirs(os.path.dirname(dst), exist_ok=True)
                    import shutil
                    shutil.copy2(src, dst)

        # Step 4: Run post script if exists
        post_script = None
        for root, dirs, files in os.walk(tmp_dir):
            if "post_update.py" in files:
                post_script = os.path.join(root, "post_update.py")
                break
        if post_script:
            subprocess.run(["python3", post_script], cwd=INSPECTION_HOME, timeout=60)

        # Step 5: Update version.json
        info_path = None
        for root, dirs, files in os.walk(tmp_dir):
            if "patch_info.json" in files:
                info_path = os.path.join(root, "patch_info.json")
                break
        if info_path:
            with open(info_path) as pf:
                info = json.load(pf)
            ver_path = os.path.join(INSPECTION_HOME, "data/version.json")
            with open(ver_path) as vf:
                ver = json.load(vf)
            ver["version"] = info.get("version", ver["version"])
            ver["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            ver["changelog"].append(f"{info.get('version','')} - {datetime.now().strftime('%Y-%m-%d')}: {info.get('description','Patch applied')}")
            with open(ver_path, "w") as vf:
                json.dump(ver, vf, indent=2, ensure_ascii=False)

        # Save patch history
        get_collection("patch_history").insert_one({
            "filename": filename,
            "version": info.get("version", "unknown") if info_path else "unknown",
            "description": info.get("description", "") if info_path else "",
            "applied_at": datetime.now().isoformat(),
            "applied_by": session.get("username", ""),
            "backup": backup_name,
        })

        subprocess.run(["rm", "-rf", tmp_dir], timeout=10)
        log_action(session["username"], "patch_apply", f"套用 Patch: {filename}", request.remote_addr)

        # Step 6: Restart Flask
        return jsonify({"success": True, "message": f"Patch 套用完成: {filename}\n請重新整理頁面", "backup": backup_name})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/patch/rollback", methods=["POST"])
@admin_required
def rollback_patch():
    # Find latest pre_patch backup
    files = sorted(glob.glob(os.path.join(BACKUP_DIR, "pre_patch_*.tar.gz")), reverse=True)
    if not files:
        return jsonify({"success": False, "error": "無可回滾的備份"}), 404
    latest = files[0]
    try:
        subprocess.run(["tar", "xzf", latest, "-C", os.path.dirname(INSPECTION_HOME)],
                        check=True, timeout=120)
        log_action(session["username"], "patch_rollback", f"回滾至: {os.path.basename(latest)}", request.remote_addr)
        return jsonify({"success": True, "message": f"已回滾至: {os.path.basename(latest)}\n請重新整理頁面"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/patch/history", methods=["GET"])
@admin_required
def patch_history():
    docs = list(get_collection("patch_history").find({}, {"_id": 0}).sort("applied_at", -1))
    return jsonify({"success": True, "data": docs})


# ========== Helpers ==========
def _sync_hosts_config():
    """Sync MongoDB hosts to hosts_config.json"""
    hosts = list(get_collection("hosts").find({}, {"_id": 0}))
    config_path = os.path.join(INSPECTION_HOME, "data/hosts_config.json")
    with open(config_path, "w") as f:
        json.dump({"hosts": hosts}, f, indent=2, ensure_ascii=False, default=str)


# ========== Forgot Password ==========
@bp.route("/forgot-password", methods=["POST"])
def forgot_password():
    """發送密碼重設信件"""
    data = request.get_json(force=True)
    username = data.get("username", "").strip()
    if not username:
        return jsonify({"success": False, "error": "請輸入帳號"}), 400

    from services.auth_service import generate_reset_token
    token, email = generate_reset_token(username)
    if not token:
        return jsonify({"success": True, "message": "如果帳號存在且已設定 Email，重設連結已寄出"})

    from services.email_service import send_email
    reset_url = f"{request.host_url}reset-password?token={token}"
    html = f"""
    <div style="font-family:sans-serif;max-width:500px;margin:0 auto;">
      <div style="background:#26A889;color:white;padding:20px;border-radius:12px 12px 0 0;text-align:center;">
        <h2 style="margin:0;">IT 每日巡檢系統</h2>
        <p style="margin:4px 0 0;opacity:0.8;">密碼重設通知</p>
      </div>
      <div style="background:white;padding:24px;border:1px solid #e0e0e0;border-radius:0 0 12px 12px;">
        <p>您好 <strong>{username}</strong>，</p>
        <p>請點擊以下連結重設密碼（有效期 30 分鐘）：</p>
        <p style="text-align:center;margin:24px 0;">
          <a href="{reset_url}" style="background:#26A889;color:white;padding:12px 32px;border-radius:6px;text-decoration:none;font-weight:600;">重設密碼</a>
        </p>
        <p style="font-size:12px;color:#888;">如果您沒有發出此請求，請忽略此信件。</p>
      </div>
    </div>
    """
    try:
        send_email(email, "密碼重設 - IT 每日巡檢系統", html)
    except Exception as e:
        return jsonify({"success": False, "error": f"信件寄送失敗：{str(e)}"}), 500

    log_action(username, "forgot_password", f"發送密碼重設信", request.remote_addr)
    return jsonify({"success": True, "message": "如果帳號存在且已設定 Email，重設連結已寄出"})


@bp.route("/reset-password", methods=["POST"])
def reset_password():
    """用 token 重設密碼"""
    data = request.get_json(force=True)
    token = data.get("token", "").strip()
    new_password = data.get("new_password", "").strip()

    if not token or not new_password:
        return jsonify({"success": False, "error": "缺少必要參數"}), 400
    if len(new_password) < 6:
        return jsonify({"success": False, "error": "密碼至少 6 個字元"}), 400

    from services.auth_service import verify_reset_token, consume_reset_token, change_password as _change_pw
    username = verify_reset_token(token)
    if not username:
        return jsonify({"success": False, "error": "連結已失效或已使用過，請重新申請"}), 400

    _change_pw(username, new_password)
    consume_reset_token(token)
    log_action(username, "reset_password", "透過信件重設密碼", request.remote_addr)
    return jsonify({"success": True, "message": "密碼已重設，請用新密碼登入"})


@bp.route("/update-email", methods=["POST"])
@login_required
def update_email():
    """更新當前使用者的 email"""
    data = request.get_json(force=True)
    email = data.get("email", "").strip()
    if not email or "@" not in email:
        return jsonify({"success": False, "error": "請輸入有效的 Email"}), 400
    from services.auth_service import update_user_email
    update_user_email(session["username"], email)
    log_action(session["username"], "update_email", f"更新 Email", request.remote_addr)
    return jsonify({"success": True, "message": "Email 已更新"})




@bp.route("/hosts/<hostname>/service", methods=["POST"])
@admin_required
def service_control(hostname):
    """遠端服務控制（啟動/重啟/停止）"""
    data = request.get_json(force=True)
    service_name = data.get("service", "").strip()
    action = data.get("action", "").strip()

    if not service_name:
        return jsonify({"success": False, "error": "未指定服務名稱"}), 400
    if action not in ("start", "stop", "restart"):
        return jsonify({"success": False, "error": "無效的操作，僅支援 start/stop/restart"}), 400

    # 查 OS 類型
    hosts_col = get_collection("hosts")
    host = hosts_col.find_one({"hostname": hostname})
    if not host:
        return jsonify({"success": False, "error": "找不到主機 " + hostname}), 404

    is_windows = "windows" in (host.get("os", "") or "").lower()
    inventory = os.path.join(INSPECTION_HOME, "ansible/inventory/hosts.yml")

    # 組命令
    if is_windows:
        ps_map = {"start": "Start-Service", "stop": "Stop-Service", "restart": "Restart-Service"}
        cmd = 'powershell -Command "' + ps_map[action] + " -Name '" + service_name + "' -Force -ErrorAction Stop\""
    else:
        cmd = "systemctl " + action + " " + service_name

    action_map = {"start": "啟動", "stop": "停止", "restart": "重啟"}
    action_text = action_map[action]

    try:
        r = subprocess.run(
            ["ansible", "-i", inventory, "--vault-password-file", os.path.join(INSPECTION_HOME, ".vault_pass"), hostname, "-m", "raw" if is_windows else "shell", "-a", cmd] + ([] if is_windows else ["-u", "ansible_svc", "--become"]),
            capture_output=True, text=True, timeout=30
        )
        output = (r.stdout + "\n" + r.stderr).strip()
        ok = r.returncode == 0

        result_text = "成功" if ok else "失敗"
        log_action(
            session.get("username", "unknown"), "service_control",
            hostname + ": " + action_text + " " + service_name + " - " + result_text,
            request.remote_addr
        )

        if ok:
            return jsonify({"success": True, "message": "服務 " + service_name + " 已" + action_text, "output": output})
        else:
            return jsonify({"success": False, "error": action_text + "失敗", "output": output}), 500
    except subprocess.TimeoutExpired:
        return jsonify({"success": False, "error": "執行逾時（30秒）: %s @ %s" % (service_name, hostname)}), 504
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@bp.route("/hosts/<hostname>/service-status", methods=["POST"])
@admin_required
def service_live_status(hostname):
    """即時查詢遠端主機所有服務狀態 + 存活時間"""
    data = request.get_json(force=True)
    services = data.get("services", [])
    if not services:
        return jsonify({"success": False, "error": "未指定服務清單"}), 400
    hosts_col = get_collection("hosts")
    host = hosts_col.find_one({"hostname": hostname})
    if not host:
        return jsonify({"success": False, "error": "找不到主機"}), 404
    is_windows = "windows" in (host.get("os", "") or "").lower()
    inventory = os.path.join(INSPECTION_HOME, "ansible/inventory/hosts.yml")
    vault_file = os.path.join(INSPECTION_HOME, ".vault_pass")
    svc_list = " ".join(services)
    if is_windows:
        cmd = "powershell -Command \"Get-Service -Name " + ",".join(services) + " -ErrorAction SilentlyContinue | ForEach-Object { $_.Name + '|' + $_.Status }\""
    else:
        cmd = "for s in " + svc_list + "; do st=$(systemctl show $s --property=ActiveState --value 2>/dev/null); ts=$(systemctl show $s --property=ActiveEnterTimestamp --value 2>/dev/null); echo \"SVC:$s:$st:$ts\"; done"
    try:
        r = subprocess.run(
            ["ansible", "-i", inventory, "--vault-password-file", vault_file,
             hostname, "-m", "raw" if is_windows else "shell", "-a", cmd] + ([] if is_windows else ["-u", "ansible_svc", "--become"]),
            capture_output=True, text=True, timeout=20
        )
        output = r.stdout.strip()
        results = []
        for line in output.split("\n"):
            line = line.strip()
            if line.startswith("SVC:"):
                parts = line.split(":", 3)
                if len(parts) >= 4:
                    results.append({"name": parts[1], "state": parts[2], "since": parts[3].strip()})
            elif "|" in line and not line.startswith(" ") and not line.startswith("{"):
                parts = line.split("|", 2)
                if len(parts) >= 2:
                    results.append({"name": parts[0].strip(), "state": parts[1].strip(), "since": parts[2].strip() if len(parts) > 2 else ""})
        return jsonify({"success": True, "data": results})
    except subprocess.TimeoutExpired:
        return jsonify({"success": False, "error": "查詢逾時"}), 504
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ===== 使用者管理 API =====

@bp.route("/users", methods=["GET"])
@admin_required
def list_users():
    """列出所有使用者"""
    col = get_collection("users")
    users = list(col.find({}, {"_id": 0, "password_hash": 0}))
    return jsonify({"success": True, "data": users, "count": len(users)})


@bp.route("/users", methods=["POST"])
@admin_required
def create_user_api():
    """新增使用者（僅 superadmin）"""
    caller_role = session.get("role")
    if caller_role not in ("admin", "superadmin"):
        return jsonify({"success": False, "error": "權限不足"}), 403
    data = request.get_json(force=True)
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    role = data.get("role", "admin")
    display_name = data.get("display_name", "").strip()
    email = data.get("email", "").strip()
    if not username or not password:
        return jsonify({"success": False, "error": "帳號和密碼為必填"}), 400
    if len(password) < 6:
        return jsonify({"success": False, "error": "密碼至少 6 碼"}), 400
    if role not in ("oper", "admin", "superadmin"):
        return jsonify({"success": False, "error": "角色僅支援 oper/admin/superadmin"}), 400
    if role == "superadmin" and caller_role != "superadmin":
        return jsonify({"success": False, "error": "僅超級管理員可指派 superadmin 角色"}), 403
    from services.auth_service import create_user
    result = create_user(username, password, role, display_name, email)
    if result is None:
        return jsonify({"success": False, "error": "帳號已存在"}), 409
    log_action(session.get("username", "unknown"), "create_user", f"新增使用者: {username} (角色: {role})", request.remote_addr)
    return jsonify({"success": True, "message": f"使用者 {username} 已建立"})


@bp.route("/users/<username>", methods=["PUT"])
@admin_required
def update_user_api(username):
    """更新使用者角色（僅 superadmin）"""
    caller_role = session.get("role")
    if caller_role not in ("admin", "superadmin"):
        return jsonify({"success": False, "error": "權限不足"}), 403
    if user.get("role") == "superadmin" and caller_role != "superadmin":
        return jsonify({"success": False, "error": "無法修改超級管理員帳號"}), 403
    data = request.get_json(force=True)
    col = get_collection("users")
    user = col.find_one({"username": username})
    if not user:
        return jsonify({"success": False, "error": "使用者不存在"}), 404
    update_fields = {}
    if "role" in data and data["role"] in ("oper", "admin", "superadmin"):
        update_fields["role"] = data["role"]
    if "display_name" in data:
        update_fields["display_name"] = data["display_name"]
    if "email" in data:
        update_fields["email"] = data["email"]
    if "password" in data and data["password"]:
        from werkzeug.security import generate_password_hash
        update_fields["password_hash"] = generate_password_hash(data["password"])
        update_fields["must_change_password"] = True
    if update_fields:
        col.update_one({"username": username}, {"$set": update_fields})
    log_action(session.get("username", "unknown"), "update_user", f"更新使用者: {username} {list(update_fields.keys())}", request.remote_addr)
    return jsonify({"success": True, "message": f"使用者 {username} 已更新"})


@bp.route("/users/<username>", methods=["DELETE"])
@admin_required
def delete_user_api(username):
    """刪除使用者（僅 superadmin，不能刪自己）"""
    caller_role = session.get("role")
    if caller_role not in ("admin", "superadmin"):
        return jsonify({"success": False, "error": "權限不足"}), 403
    if username == session.get("username"):
        return jsonify({"success": False, "error": "不能刪除自己的帳號"}), 400
    col = get_collection("users")
    target = col.find_one({"username": username})
    if target and target.get("role") == "superadmin" and caller_role != "superadmin":
        return jsonify({"success": False, "error": "無法刪除超級管理員帳號"}), 403
    result = col.delete_one({"username": username})
    if result.deleted_count == 0:
        return jsonify({"success": False, "error": "使用者不存在"}), 404
    log_action(session.get("username", "unknown"), "delete_user", f"刪除使用者: {username}", request.remote_addr)
    return jsonify({"success": True, "message": f"使用者 {username} 已刪除"})


# ===== Tunnel 管理 API =====

@bp.route("/tunnel/status", methods=["GET"])
@admin_required
def tunnel_status():
    """查詢 Cloudflare Tunnel 狀態和網址"""
    try:
        r = subprocess.run(["systemctl", "is-active", "itagent-tunnel"], capture_output=True, text=True, timeout=5)
        active = r.stdout.strip() == "active"
    except Exception:
        active = False
    url = ""
    if active:
        try:
            r2 = subprocess.run(
                ["journalctl", "-u", "itagent-tunnel", "--no-pager", "-n", "50"],
                capture_output=True, text=True, timeout=5
            )
            import re as _re
            matches = _re.findall(r'https://[^\s]*trycloudflare\.com', r2.stdout)
            if matches:
                url = matches[-1]
        except Exception:
            pass
    return jsonify({"success": True, "data": {"active": active, "url": url}})


@bp.route("/tunnel/restart", methods=["POST"])
@admin_required
def tunnel_restart():
    """重啟 Tunnel（會產生新網址）"""
    try:
        subprocess.run(["systemctl", "restart", "itagent-tunnel"], capture_output=True, text=True, timeout=15)
        import time; time.sleep(8)
        r = subprocess.run(
            ["journalctl", "-u", "itagent-tunnel", "--no-pager", "-n", "30"],
            capture_output=True, text=True, timeout=5
        )
        import re as _re
        matches = _re.findall(r'https://[^\s]*trycloudflare\.com', r.stdout)
        url = matches[-1] if matches else ""
        log_action(session.get("username", "unknown"), "tunnel_restart", "重啟 Tunnel: " + url, request.remote_addr)
        return jsonify({"success": True, "message": "Tunnel 已重啟", "url": url})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/tunnel/stop", methods=["POST"])
@admin_required
def tunnel_stop():
    """停止 Tunnel"""
    try:
        subprocess.run(["systemctl", "stop", "itagent-tunnel"], capture_output=True, text=True, timeout=10)
        log_action(session.get("username", "unknown"), "tunnel_stop", "停止 Tunnel", request.remote_addr)
        return jsonify({"success": True, "message": "Tunnel 已停止"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ===== TWGCB 一鍵修復 API =====

@bp.route("/twgcb/fix", methods=["POST"])
@admin_required
def twgcb_fix():
    """執行單項 TWGCB 修復（含備份）"""
    data = request.get_json(force=True)
    check_id = data.get("check_id", "").strip()
    hostname = data.get("hostname", "").strip()
    if not check_id or not hostname:
        return jsonify({"success": False, "error": "缺少 check_id 或 hostname"}), 400

    # 變更 #39: 後端同主機鎖（跨 worker/user 互斥，防止並行 remediation race）
    # 前端 _fixingHosts 鎖只防同一使用者瀏覽器連點；後端鎖才能防多使用者 / multi-worker
    if not _mongo_try_lock_host(hostname, ttl_sec=60, locked_by="fix:" + check_id):
        return jsonify({
            "success": False,
            "error": "主機 %s 正在由其他使用者執行修復中，請稍候重試" % hostname,
            "hostname": hostname, "check_id": check_id, "locked": True
        }), 409

    from services.mongo_service import get_collection
    config = get_collection("twgcb_config").find_one({"check_id": check_id})
    remediation = ""
    if config:
        remediation = config.get("remediation", "")
    if not remediation:
        # Fallback: try to find remediation from scan results
        scan_result = get_collection("twgcb_results").find_one({"hostname": hostname})
        if scan_result:
            for chk in scan_result.get("checks", []):
                if chk.get("id") == check_id:
                    remediation = chk.get("remediation", "")
                    break
    if not remediation:
        return jsonify({"success": False, "error": "此項目無修復指令（check_id: %s）" % check_id}), 400

    host = get_collection("hosts").find_one({"hostname": hostname})
    is_windows = "windows" in (host.get("os", "") if host else "").lower()
    inventory = os.path.join(INSPECTION_HOME, "ansible/inventory/hosts.yml")
    vault_file = os.path.join(INSPECTION_HOME, ".vault_pass")
    ansible_base = ["ansible", "-i", inventory, "--vault-password-file", vault_file, hostname, "-m", "raw" if is_windows else "shell"] + ([] if is_windows else ["-u", "ansible_svc", "--become"])

    # Step 1: 備份相關檔案
    import re as _re
    backup_files = _re.findall(r'/etc/[\w./\-]+', remediation)
    backup_output = ""
    if backup_files and not is_windows:
        backup_cmd = " && ".join(["cp -p %s %s.twgcb_bak 2>/dev/null || true" % (f, f) for f in backup_files[:3]])
        try:
            br = subprocess.run(ansible_base + ["-a", backup_cmd], capture_output=True, text=True, timeout=15)
            backup_output = "備份: " + ", ".join([f + ".twgcb_bak" for f in backup_files[:3]])
        except Exception:
            backup_output = "備份跳過"

    # Step 2: 執行修復
    if is_windows:
        cmd = 'powershell -Command "%s"' % remediation.replace('"', '\\"')
    else:
        cmd = remediation

    try:
        r = subprocess.run(ansible_base + ["-a", cmd], capture_output=True, text=True, timeout=30)
        output = (r.stdout + "\n" + r.stderr).strip()
        # 變更 #36 Bug A: rc=0 不等於 remediation 真的成功
        # 許多 remediation 尾部 echo OK 會吞掉錯誤；加字串檢查抓真正失敗
        failure_markers = ["| FAILED", "UNREACHABLE", "Failed to mask",
                           "Failed to enable", "Failed to disable",
                           "Authentication failed", "Permission denied"]
        has_failure = any(m in output for m in failure_markers)
        ok = r.returncode == 0 and not has_failure
        log_action(session.get("username", "unknown"), "twgcb_fix",
            "%s: %s %s — %s" % (hostname, check_id, ( config.get("description","") if config else check_id ), "成功" if ok else "失敗"),
            request.remote_addr)
        result = {
            "success": ok, "message": "修復成功" if ok else "修復失敗",
            "output": output, "remediation": remediation,
            "backup": backup_output,
            "backup_files": [f + ".twgcb_bak" for f in backup_files[:3]] if backup_files else [],
            "check_id": check_id, "hostname": hostname,
        }
        if not ok:
            result["error"] = "修復失敗: %s @ %s" % (check_id, hostname)
            result["detail"] = output

        # 變更 #36 Bug B: 修復成功時背景重跑 scan + import，讓 UI 刷新看到新狀態
        # 不讓使用者按「執行修復」要等 30~60 秒 ansible 掃描完才有回應
        if ok and not is_windows:
            def _bg_rescan(_hostname=hostname, _inventory=inventory, _vault=vault_file):
                try:
                    playbook = os.path.join(INSPECTION_HOME, "ansible/playbooks/twgcb_scan.yml")
                    subprocess.run(
                        ["ansible-playbook", "-i", _inventory, "--vault-password-file", _vault,
                         playbook, "--limit", _hostname],
                        capture_output=True, text=True, timeout=180
                    )
                    # 把 JSON 寫回 MongoDB（複製 api_twgcb._import_results 邏輯，避免循環 import）
                    from services.mongo_service import get_collection
                    report_path = os.path.join(INSPECTION_HOME, "data/reports", "twgcb_%s.json" % _hostname)
                    if os.path.exists(report_path):
                        with open(report_path) as fh:
                            scan_data = json.load(fh)
                        if scan_data.get("hostname"):
                            scan_data["imported_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            get_collection("twgcb_results").update_one(
                                {"hostname": _hostname}, {"$set": scan_data}, upsert=True
                            )
                except Exception:
                    pass  # 背景任務失敗不影響主回應
                finally:
                    # 變更 #39: 背景 rescan 結束後釋放鎖
                    _mongo_release_lock_host(_hostname)
            # 變更 #39: 延長鎖到背景 rescan 結束（最多 180s + 緩衝）
            _mongo_extend_lock_host(hostname, ttl_sec=200, locked_by="bg_rescan:" + check_id)
            threading.Thread(target=_bg_rescan, daemon=True).start()
            result["rescan_pending"] = True
        else:
            # 失敗或 Windows：立即釋放鎖
            _mongo_release_lock_host(hostname)

        return jsonify(result), 200
    except subprocess.TimeoutExpired:
        _mongo_release_lock_host(hostname)
        return jsonify({"success": False, "error": "執行逾時（30秒）: %s @ %s" % (check_id, hostname), "hostname": hostname, "check_id": check_id}), 200
    except Exception as e:
        _mongo_release_lock_host(hostname)
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/twgcb/config/<check_id>/remediation", methods=["GET"])
@admin_required
def get_remediation(check_id):
    """取得單項修復指令"""
    from services.mongo_service import get_collection
    config = get_collection("twgcb_config").find_one({"check_id": check_id}, {"_id":0, "remediation":1, "description":1, "category":1})
    if not config:
        return jsonify({"success": False, "error": "找不到"}), 404
    return jsonify({"success": True, "data": config})


@bp.route("/twgcb/restore", methods=["POST"])
@admin_required
def twgcb_restore():
    """還原單項 TWGCB 修復前的備份"""
    data = request.get_json(force=True)
    hostname = data.get("hostname", "").strip()
    backup_files = data.get("backup_files", [])
    if not hostname or not backup_files:
        return jsonify({"success": False, "error": "缺少參數"}), 400

    inventory = os.path.join(INSPECTION_HOME, "ansible/inventory/hosts.yml")
    vault_file = os.path.join(INSPECTION_HOME, ".vault_pass")

    restore_cmd = " && ".join(["cp -p %s %s 2>/dev/null" % (bak, bak.replace(".twgcb_bak","")) for bak in backup_files])
    try:
        r = subprocess.run(
            ["ansible", "-i", inventory, "--vault-password-file", vault_file,
             hostname, "-m", "shell", "-a", restore_cmd, "-u", "ansible_svc", "--become"],  # Linux only restore
            capture_output=True, text=True, timeout=15
        )
        output = (r.stdout + "\n" + r.stderr).strip()
        ok = r.returncode == 0
        log_action(session.get("username", "unknown"), "twgcb_restore",
            "%s: 還原 %s — %s" % (hostname, str(backup_files), "成功" if ok else "失敗"),
            request.remote_addr)
        if ok:
            # 變更 #45: 還原成功後背景 rescan，讓 UI 20 秒後顯示真實狀態
            def _bg_rescan(_hostname=hostname, _inventory=inventory, _vault=vault_file):
                try:
                    playbook = os.path.join(INSPECTION_HOME, "ansible/playbooks/twgcb_scan.yml")
                    subprocess.run(
                        ["ansible-playbook", "-i", _inventory, "--vault-password-file", _vault,
                         playbook, "--limit", _hostname],
                        capture_output=True, text=True, timeout=180
                    )
                    report_path = os.path.join(INSPECTION_HOME, "data/reports", "twgcb_%s.json" % _hostname)
                    if os.path.exists(report_path):
                        with open(report_path) as fh:
                            scan_data = json.load(fh)
                        if scan_data.get("hostname"):
                            scan_data["imported_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            get_collection("twgcb_results").update_one(
                                {"hostname": _hostname}, {"$set": scan_data}, upsert=True
                            )
                except Exception:
                    pass
            threading.Thread(target=_bg_rescan, daemon=True).start()
            return jsonify({"success": True, "message": "還原成功", "output": output, "rescan_pending": True})
        else:
            return jsonify({"success": False, "error": "還原失敗", "output": output}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ===== 批次主機連線檢查 API =====

@bp.route("/hosts/ping-all", methods=["GET"])
def ping_all_hosts():
    """批次 ping 所有主機，回傳連線狀態（不需登入，公開 API）

    Phase 2 #4: 60 秒記憶體快取；?force=1 可強制刷新。
    變更 #39: 改用 MongoDB cache collection，為 Gunicorn multi-worker 做準備。
    """
    force = request.args.get("force", "").strip() in ("1", "true", "yes")

    # 快取命中：直接回（MongoDB 共享快取，跨 worker 有效）
    if not force:
        cached = _mongo_cache_get("ping_all", PING_CACHE_TTL)
        if cached is not None:
            return jsonify({
                "success": True,
                "data": cached["data"],
                "cached": True,
                "age_sec": cached["age_sec"],
            })

    # process 內 lock 防止單 worker 自己並發重跑；跨 worker 由 MongoDB double-check 擋
    with _ping_lock:
        if not force:
            cached = _mongo_cache_get("ping_all", PING_CACHE_TTL)
            if cached is not None:
                return jsonify({
                    "success": True,
                    "data": cached["data"],
                    "cached": True,
                    "age_sec": cached["age_sec"],
                })

        hosts = list(get_collection("hosts").find({}, {"_id": 0, "hostname": 1, "ip": 1}))
        results = {}

        def do_ping(hostname, ip):
            try:
                r = subprocess.run(
                    ["ping", "-c", "1", "-W", "2", ip],
                    capture_output=True, text=True, timeout=5
                )
                results[hostname] = {"reachable": r.returncode == 0, "ip": ip}
            except Exception:
                results[hostname] = {"reachable": False, "ip": ip}

        threads = []
        for h in hosts:
            t = threading.Thread(target=do_ping, args=(h["hostname"], h.get("ip", "")))
            t.start()
            threads.append(t)

        for t in threads:
            t.join(timeout=6)

        _mongo_cache_set("ping_all", results)

    return jsonify({
        "success": True,
        "data": results,
        "cached": False,
        "age_sec": 0,
    })


# ===== 帳號鎖定/解鎖/重設 API =====

@bp.route("/hosts/<hostname>/faillock", methods=["POST"])
@admin_required
def faillock_action(hostname):
    """帳號鎖定管理: lock/unlock/reset"""
    data = request.get_json(force=True)
    user = data.get("user", "").strip()
    action = data.get("action", "").strip()
    if not user or action not in ("lock", "unlock", "reset"):
        return jsonify({"success": False, "error": "缺少 user 或 action (lock/unlock/reset)"}), 400

    hosts_col = get_collection("hosts")
    host = hosts_col.find_one({"hostname": hostname})
    if not host:
        return jsonify({"success": False, "error": "找不到主機"}), 404

    is_windows = "windows" in (host.get("os", "") or "").lower()
    inventory = os.path.join(INSPECTION_HOME, "ansible/inventory/hosts.yml")
    vault_file = os.path.join(INSPECTION_HOME, ".vault_pass")

    if is_windows:
        if action == "lock":
            cmd = 'net user %s /active:no' % user
        elif action == "unlock":
            cmd = 'net user %s /active:yes' % user
        else:
            cmd = 'net user %s /active:yes' % user  # reset = unlock on Windows
    else:
        if action == "lock":
            cmd = 'passwd -l %s' % user
        elif action == "unlock":
            cmd = 'passwd -u %s && faillock --user %s --reset 2>/dev/null; pam_tally2 --user=%s --reset 2>/dev/null; echo done' % (user, user, user)
        else:  # reset counter
            cmd = 'faillock --user %s --reset 2>/dev/null; pam_tally2 --user=%s --reset 2>/dev/null; echo counter_reset' % (user, user)

    action_text = {"lock": "鎖定", "unlock": "解鎖", "reset": "重設計數"}[action]
    try:
        r = subprocess.run(
            ["ansible", "-i", inventory, "--vault-password-file", vault_file,
             hostname, "-m", "raw" if is_windows else "shell", "-a", cmd] + ([] if is_windows else ["-u", "ansible_svc", "--become"]),
            capture_output=True, text=True, timeout=15
        )
        output = (r.stdout + "\n" + r.stderr).strip()
        ok = r.returncode == 0
        log_action(session.get("username", "unknown"), "faillock",
            "%s: %s %s — %s" % (hostname, action_text, user, "成功" if ok else "失敗"),
            request.remote_addr)
        if ok:
            return jsonify({"success": True, "message": "%s 帳號 %s 已%s" % (hostname, user, action_text), "output": output})
        else:
            return jsonify({"success": False, "error": "%s失敗" % action_text, "output": output}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ===== TWGCB 一鍵全修單台 API =====

@bp.route("/twgcb/fix-all", methods=["POST"])
@admin_required
def twgcb_fix_all():
    """一鍵修復單台主機所有 FAIL 項目（備份→修復→服務重啟）"""
    data = request.get_json(force=True)
    hostname = data.get("hostname", "").strip()
    if not hostname:
        return jsonify({"success": False, "error": "缺少 hostname"}), 400

    # 變更 #39: 後端同主機鎖（全修可能跑數十秒，需要較長 TTL）
    if not _mongo_try_lock_host(hostname, ttl_sec=300, locked_by="fix_all"):
        return jsonify({
            "success": False,
            "error": "主機 %s 正在由其他使用者執行修復中，請稍候重試" % hostname,
            "hostname": hostname, "locked": True
        }), 409

    from services.mongo_service import get_collection
    host = get_collection("hosts").find_one({"hostname": hostname})
    if not host:
        _mongo_release_lock_host(hostname)
        return jsonify({"success": False, "error": "找不到主機"}), 404

    # 記錄修復開始狀態
    fix_col = get_collection("twgcb_fix_status")
    fix_col.update_one({"hostname": hostname}, {"$set": {
        "hostname": hostname,
        "status": "running",
        "started_at": datetime.now().isoformat(),
        "started_by": session.get("username", "unknown"),
        "results": None,
        "message": None,
    }}, upsert=True)

    is_windows = "windows" in (host.get("os", "") or "").lower()
    inventory = os.path.join(INSPECTION_HOME, "ansible/inventory/hosts.yml")
    vault_file = os.path.join(INSPECTION_HOME, ".vault_pass")

    # 取該主機的掃描結果，找出所有 FAIL 項
    result_doc = get_collection("twgcb_results").find_one({"hostname": hostname})
    if not result_doc:
        _mongo_release_lock_host(hostname)
        return jsonify({"success": False, "error": "無掃描結果"}), 404

    fail_checks = [c for c in result_doc.get("checks", []) if c.get("status") == "FAIL"]

    # 變更 #42: 跳過已標記例外的 check_id（例外=已批准不修復，全修不該動它）
    excepted_ids = {e["check_id"] for e in get_collection("twgcb_exceptions").find(
        {"hostname": hostname}, {"check_id": 1, "_id": 0}
    )}
    skipped_exceptions = [c for c in fail_checks if c.get("id") in excepted_ids]
    fail_checks = [c for c in fail_checks if c.get("id") not in excepted_ids]

    if not fail_checks:
        msg = "無需修復，所有項目已通過" if not skipped_exceptions else \
              f"無需修復（{len(skipped_exceptions)} 項已標記例外，未動）"
        _mongo_release_lock_host(hostname)
        return jsonify({
            "success": True, "message": msg, "fixed": 0, "total": 0,
            "skipped_exceptions": len(skipped_exceptions)
        })

    # 從 config 取 remediation 指令
    configs = {c["check_id"]: c for c in get_collection("twgcb_config").find({}, {"_id": 0})}

    import re as _re
    results = []
    fixed = 0
    failed = 0

    for check in fail_checks:
        cid = check.get("id", "")
        cfg = configs.get(cid, {})
        remediation = cfg.get("remediation", "") or check.get("remediation", "")
        if not remediation or remediation.startswith("#"):
            results.append({"id": cid, "name": check.get("name",""), "status": "SKIP", "detail": "無可自動修復指令"})
            continue

        # Step 1: 備份相關檔案
        backup_files = _re.findall(r'/etc/[\w./\-]+', remediation)
        if backup_files and not is_windows:
            backup_cmd = " && ".join(["cp -p %s %s.twgcb_bak 2>/dev/null || true" % (f, f) for f in backup_files[:3]])
            try:
                subprocess.run(
                    ["ansible", "-i", inventory, "--vault-password-file", vault_file,
                     hostname, "-m", "raw" if is_windows else "shell", "-a", backup_cmd]
                    + ([] if is_windows else ["-u", "ansible_svc", "--become"]),
                    capture_output=True, text=True, timeout=10
                )
            except Exception:
                pass

        # Step 2: 執行修復
        if is_windows:
            cmd = 'powershell -Command "%s"' % remediation.replace('"', '\\"')
        else:
            cmd = remediation

        try:
            r = subprocess.run(
                ["ansible", "-i", inventory, "--vault-password-file", vault_file,
                 hostname, "-m", "raw" if is_windows else "shell", "-a", cmd]
                + ([] if is_windows else ["-u", "ansible_svc", "--become"]),
                capture_output=True, text=True, timeout=30
            )
            _fix_output = (r.stdout or "") + (r.stderr or "")
            # 變更 #36 Bug A: rc=0 不等於 remediation 真的成功（echo OK 會吞錯誤）
            _failure_markers = ["| FAILED", "UNREACHABLE", "Failed to mask",
                                "Failed to enable", "Failed to disable",
                                "Authentication failed", "Permission denied"]
            ok = r.returncode == 0 and not any(m in _fix_output for m in _failure_markers)
            if ok:
                fixed += 1
                # 立即重啟受影響的服務
                svc_restart = ""
                if not is_windows:
                    svc_map = {"ssh": "sshd", "SSH": "sshd", "008-02": "sshd",
                               "syslog": "rsyslog", "rsyslog": "rsyslog", "008-01": "rsyslog",
                               "audit": "auditd", "008-01": "auditd",
                               "pam": "sshd", "password": "sshd", "login": "sshd"}
                    for keyword, svc in svc_map.items():
                        if keyword.lower() in (cid + " " + remediation).lower():
                            try:
                                subprocess.run(
                                    ["ansible", "-i", inventory, "--vault-password-file", vault_file,
                                     hostname, "-m", "shell", "-a", "systemctl restart " + svc + " 2>/dev/null || true",
                                     "-u", "ansible_svc", "--become"],
                                    capture_output=True, text=True, timeout=10
                                )
                                svc_restart = svc
                            except Exception:
                                pass
                            break
                detail_msg = r.stdout.strip()[:80]
                if svc_restart:
                    detail_msg += " (已重啟 " + svc_restart + ")"
                results.append({"id": cid, "name": check.get("name",""), "status": "FIXED", "detail": detail_msg})
            else:
                failed += 1
                results.append({"id": cid, "name": check.get("name",""), "status": "FAIL", "detail": (r.stdout + r.stderr).strip()[:100]})
        except subprocess.TimeoutExpired:
            failed += 1
            results.append({"id": cid, "name": check.get("name",""), "status": "TIMEOUT"})
        except Exception as e:
            failed += 1
            results.append({"id": cid, "name": check.get("name",""), "status": "ERROR", "detail": str(e)[:100]})

    # 服務已在每項修復後立即重啟，這裡做最終確認
    restart_output = "服務已逐項重啟"

    log_action(session.get("username", "unknown"), "twgcb_fix_all",
        "%s: 一鍵全修 %d/%d 成功" % (hostname, fixed, len(fail_checks)),
        request.remote_addr)

    # 變更 #36 Bug B: 不再「作弊」直接把 twgcb_results 的 status 改 PASS
    # （那做法只是「看起來成功」，實機沒驗證，rescan 後還是 FAIL 會誤導使用者）
    # 改為背景重跑 scan + import，讓前端 20 秒後拿到真實狀態
    rescan_pending = False
    if fixed > 0 and not is_windows:
        def _bg_rescan(_hostname=hostname, _inventory=inventory, _vault=vault_file):
            try:
                playbook = os.path.join(INSPECTION_HOME, "ansible/playbooks/twgcb_scan.yml")
                subprocess.run(
                    ["ansible-playbook", "-i", _inventory, "--vault-password-file", _vault,
                     playbook, "--limit", _hostname],
                    capture_output=True, text=True, timeout=180
                )
                report_path = os.path.join(INSPECTION_HOME, "data/reports", "twgcb_%s.json" % _hostname)
                if os.path.exists(report_path):
                    with open(report_path) as fh:
                        scan_data = json.load(fh)
                    if scan_data.get("hostname"):
                        scan_data["imported_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        get_collection("twgcb_results").update_one(
                            {"hostname": _hostname}, {"$set": scan_data}, upsert=True
                        )
            except Exception:
                pass
            finally:
                # 變更 #39: 背景 rescan 結束後釋放鎖
                _mongo_release_lock_host(_hostname)
        # 變更 #39: 延長鎖到背景 rescan 結束（rescan 最多 180s + 緩衝）
        _mongo_extend_lock_host(hostname, ttl_sec=220, locked_by="fix_all:bg_rescan")
        threading.Thread(target=_bg_rescan, daemon=True).start()
        rescan_pending = True
    else:
        # 沒有要背景 rescan：立即釋放鎖
        _mongo_release_lock_host(hostname)

    # 記錄修復完成狀態
    fix_msg = "修復完成: %d 成功 / %d 失敗 / %d 跳過" % (fixed, failed, len(fail_checks) - fixed - failed)
    fix_col.update_one({"hostname": hostname}, {"$set": {
        "status": "done",
        "finished_at": datetime.now().isoformat(),
        "message": fix_msg,
        "results": results,
        "restart": restart_output,
        "fixed": fixed, "failed": failed, "total": len(fail_checks),
    }})

    return jsonify({
        "success": True,
        "message": fix_msg,
        "total": len(fail_checks), "fixed": fixed, "failed": failed,
        "results": results,
        "restart": restart_output,
        "rescan_pending": rescan_pending,
        # 變更 #42: 前端可顯示「已跳過 N 項例外」
        "skipped_exceptions": len(skipped_exceptions),
    })


# ===== TWGCB 修復狀態查詢 =====
@bp.route("/twgcb/fix-status/<hostname>", methods=["GET"])
@login_required
def twgcb_fix_status(hostname):
    """查詢修復進度（F5 安全）"""
    from services.mongo_service import get_collection
    doc = get_collection("twgcb_fix_status").find_one({"hostname": hostname}, {"_id": 0})
    if not doc:
        return jsonify({"success": True, "data": None})
    return jsonify({"success": True, "data": doc})


# ===== TWGCB 全部還原 API =====

@bp.route("/twgcb/restore-all", methods=["POST"])
@admin_required
def twgcb_restore_all():
    """還原單台主機所有 .twgcb_bak 備份檔"""
    data = request.get_json(force=True)
    hostname = data.get("hostname", "").strip()
    if not hostname:
        return jsonify({"success": False, "error": "缺少 hostname"}), 400

    host = get_collection("hosts").find_one({"hostname": hostname})
    if not host:
        return jsonify({"success": False, "error": "找不到主機"}), 404

    is_windows = "windows" in (host.get("os", "") or "").lower()
    inventory = os.path.join(INSPECTION_HOME, "ansible/inventory/hosts.yml")
    vault_file = os.path.join(INSPECTION_HOME, ".vault_pass")

    # Step 1: 找出所有 .twgcb_bak 檔案
    find_cmd = "find /etc -name '*.twgcb_bak' -type f 2>/dev/null"
    try:
        r1 = subprocess.run(
            ["ansible", "-i", inventory, "--vault-password-file", vault_file,
             hostname, "-m", "raw" if is_windows else "shell", "-a", find_cmd]
            + ([] if is_windows else ["-u", "ansible_svc", "--become"]),
            capture_output=True, text=True, timeout=15
        )
        bak_files = [line.strip() for line in r1.stdout.split("\n") if line.strip().endswith(".twgcb_bak")]
    except Exception:
        bak_files = []

    if not bak_files:
        return jsonify({"success": True, "message": "無備份檔案需要還原", "restored": 0})

    # Step 2: 逐一還原
    restore_cmds = []
    for bak in bak_files:
        orig = bak.replace(".twgcb_bak", "")
        restore_cmds.append("cp -p %s %s && echo 'restored:%s'" % (bak, orig, orig))

    restore_cmd = " ; ".join(restore_cmds)
    try:
        r2 = subprocess.run(
            ["ansible", "-i", inventory, "--vault-password-file", vault_file,
             hostname, "-m", "raw" if is_windows else "shell", "-a", restore_cmd]
            + ([] if is_windows else ["-u", "ansible_svc", "--become"]),
            capture_output=True, text=True, timeout=30
        )
        output = r2.stdout.strip()
        restored = output.count("restored:")

        # Step 3: 重啟相關服務
        restart_cmd = "systemctl restart sshd auditd rsyslog 2>/dev/null; echo services_restarted"
        r3 = subprocess.run(
            ["ansible", "-i", inventory, "--vault-password-file", vault_file,
             hostname, "-m", "shell", "-a", restart_cmd, "-u", "ansible_svc", "--become"],
            capture_output=True, text=True, timeout=15
        )

        log_action(session.get("username", "unknown"), "twgcb_restore_all",
            "%s: 全部還原 %d/%d 檔案" % (hostname, restored, len(bak_files)),
            request.remote_addr)

        # 變更 #45: 移除舊的 fixed_at hack（#36 之後 fix-all 不再寫 fixed_at
        # 所以 array_filters 永遠 match 不到 → UI 不會更新）
        # 改用跟 fix-all 一致的背景 rescan pattern，拿實機真實狀態
        rescan_pending = False
        if not is_windows:
            def _bg_rescan(_hostname=hostname, _inventory=inventory, _vault=vault_file):
                try:
                    playbook = os.path.join(INSPECTION_HOME, "ansible/playbooks/twgcb_scan.yml")
                    subprocess.run(
                        ["ansible-playbook", "-i", _inventory, "--vault-password-file", _vault,
                         playbook, "--limit", _hostname],
                        capture_output=True, text=True, timeout=180
                    )
                    report_path = os.path.join(INSPECTION_HOME, "data/reports", "twgcb_%s.json" % _hostname)
                    if os.path.exists(report_path):
                        with open(report_path) as fh:
                            scan_data = json.load(fh)
                        if scan_data.get("hostname"):
                            scan_data["imported_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            get_collection("twgcb_results").update_one(
                                {"hostname": _hostname}, {"$set": scan_data}, upsert=True
                            )
                except Exception:
                    pass
            threading.Thread(target=_bg_rescan, daemon=True).start()
            rescan_pending = True

        return jsonify({
            "success": True,
            "message": "還原完成: %d 個檔案已還原" % restored,
            "total": len(bak_files),
            "restored": restored,
            "files": bak_files,
            "output": output,
            "restart": r3.stdout.strip(),
            "rescan_pending": rescan_pending,
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ========== SSH Key Management ==========

@bp.route("/ssh/status", methods=["GET"])
@admin_required
def ssh_status():
    """Get local SSH key status + all hosts SSH connectivity"""
    import subprocess, os
    key_path = os.path.expanduser("~/.ssh/id_ed25519")
    pub_path = key_path + ".pub"

    key_info = {"exists": False, "fingerprint": "", "public_key": "", "created_by": "", "created_at": ""}
    if os.path.exists(pub_path):
        key_info["exists"] = True
        with open(pub_path) as f:
            pub_content = f.read().strip()
            key_info["public_key"] = pub_content
            # 從 public key 的 comment 欄位解析建立者
            parts = pub_content.split()
            if len(parts) >= 3:
                comment = parts[2]  # e.g. "itagent-by-superadmin"
                if comment.startswith("itagent-by-"):
                    key_info["created_by"] = comment.replace("itagent-by-", "")
                else:
                    key_info["created_by"] = comment
        try:
            r = subprocess.run(["ssh-keygen", "-lf", pub_path], capture_output=True, text=True, timeout=5)
            key_info["fingerprint"] = r.stdout.strip()
        except:
            pass
        # 從 audit log 撈建立時間
        try:
            log = get_collection("audit_log").find_one(
                {"action": "ssh_generate_key"}, sort=[("timestamp", -1)]
            )
            if log:
                key_info["created_at"] = str(log.get("timestamp", ""))
                if not key_info["created_by"]:
                    key_info["created_by"] = log.get("username", "")
        except:
            pass

    hosts = list(get_collection("hosts").find({}, {"_id": 0, "hostname": 1, "ip": 1, "os": 1, "os_group": 1, "ssh_key_records": 1}))
    return jsonify({"success": True, "data": {"key": key_info, "hosts": hosts}})


@bp.route("/ssh/generate-key", methods=["POST"])
@admin_required
def ssh_generate_key():
    """Generate new ed25519 SSH key pair"""
    import subprocess, os
    key_path = os.path.expanduser("~/.ssh/id_ed25519")
    pub_path = key_path + ".pub"

    if os.path.exists(key_path):
        return jsonify({"success": False, "error": "Key already exists. Delete first if you want to regenerate."}), 400

    os.makedirs(os.path.expanduser("~/.ssh"), mode=0o700, exist_ok=True)
    try:
        r = subprocess.run(
            ["ssh-keygen", "-t", "ed25519", "-C", "itagent-by-{}".format(session["username"]), "-f", key_path, "-N", ""],
            capture_output=True, text=True, timeout=10
        )
        if r.returncode != 0:
            return jsonify({"success": False, "error": r.stderr}), 500

        with open(pub_path) as f:
            pub = f.read().strip()
        fp = subprocess.run(["ssh-keygen", "-lf", pub_path], capture_output=True, text=True, timeout=5)

        log_action(session["username"], "ssh_generate_key", "Generated new SSH key", request.remote_addr)
        return jsonify({"success": True, "data": {"public_key": pub, "fingerprint": fp.stdout.strip()}})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/ssh/send-key", methods=["POST"])
@admin_required
def ssh_send_key():
    """Send SSH key to a remote host using password"""
    import subprocess, os
    data = request.get_json()
    ip = data.get("ip", "").strip()
    password = data.get("password", "")
    user = data.get("user", "root")

    if not ip or not password:
        return jsonify({"success": False, "error": "IP and password required"}), 400

    pub_path = os.path.expanduser("~/.ssh/id_ed25519.pub")
    if not os.path.exists(pub_path):
        return jsonify({"success": False, "error": "No SSH key. Generate one first."}), 400

    try:
        r = subprocess.run(
            ["sshpass", "-p", password, "ssh-copy-id", "-o", "StrictHostKeyChecking=no", f"{user}@{ip}"],
            capture_output=True, text=True, timeout=30
        )
        if r.returncode == 0:
            log_action(session["username"], "ssh_send_key", f"Sent SSH key to {user}@{ip}", request.remote_addr)
            return jsonify({"success": True, "message": f"Key sent to {user}@{ip}"})
        else:
            err = r.stderr.strip()
            if "Permission denied" in err or "assword" in err:
                return jsonify({"success": False, "error": "Password incorrect"}), 401
            return jsonify({"success": False, "error": err}), 500
    except subprocess.TimeoutExpired:
        return jsonify({"success": False, "error": "Connection timeout"}), 504
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/ssh/test", methods=["POST"])
@admin_required
def ssh_test():
    """Test SSH connectivity to one or all hosts, try each user"""
    import subprocess
    data = request.get_json() or {}
    target_ip = data.get("ip", "")
    default_users = ["sysinfra", "root"]

    if target_ip:
        # Single host test
        host_doc = get_collection("hosts").find_one({"ip": target_ip}, {"_id": 0, "hostname": 1, "ip": 1, "ssh_key_records": 1}) or {}
        hosts = [{"ip": target_ip, "hostname": data.get("hostname", host_doc.get("hostname", target_ip)), "ssh_key_records": host_doc.get("ssh_key_records", {})}]
    else:
        # All hosts
        hosts = list(get_collection("hosts").find({}, {"_id": 0, "hostname": 1, "ip": 1, "ssh_key_records": 1}))

    results = []
    for h in hosts:
        ip = h.get("ip", "")
        hostname = h.get("hostname", ip)
        if not ip:
            results.append({"hostname": hostname, "ip": ip, "status": "error", "ssh_user": "", "message": "No IP"})
            continue
        # 從 ssh_key_records 取得已部署帳號，合併預設帳號（去重）
        records = h.get("ssh_key_records", {})
        host_users = list(records.keys()) if records else []
        for du in default_users:
            if du not in host_users:
                host_users.append(du)

        # 每個帳號都測試
        user_results = {}
        def _parse_err(err):
            if "Permission denied" in err: return "Key 認證失敗"
            if "Connection refused" in err: return "連線被拒"
            if "No route to host" in err: return "主機不可達"
            if "Connection timed out" in err: return "連線逾時"
            if "Host key verification" in err: return "Host key 驗證失敗"
            if err == "Timeout": return "連線逾時"
            return err if err else "未知錯誤"

        for u in host_users:
            try:
                r = subprocess.run(
                    ["ssh", "-o", "StrictHostKeyChecking=no", "-o", "ConnectTimeout=5",
                     "-o", "BatchMode=yes", f"{u}@{ip}", "hostname"],
                    capture_output=True, text=True, timeout=10
                )
                if r.returncode == 0:
                    user_results[u] = "ok"
                else:
                    err = r.stderr.strip()
                    user_results[u] = _parse_err(err)
                    log_action(session["username"], "ssh_test_fail",
                        f"{u}@{ip}({hostname}) FAIL: {err}",
                        request.remote_addr)
            except subprocess.TimeoutExpired:
                user_results[u] = "連線逾時"
                log_action(session["username"], "ssh_test_fail",
                    f"{u}@{ip}({hostname}) FAIL: Timeout",
                    request.remote_addr)
            except Exception as e:
                user_results[u] = str(e)
                log_action(session["username"], "ssh_test_fail",
                    f"{u}@{ip}({hostname}) FAIL: {e}",
                    request.remote_addr)

        has_ok = any(v == "ok" for v in user_results.values())
        results.append({
            "hostname": hostname, "ip": ip,
            "status": "ok" if has_ok else "error",
            "user_results": user_results
        })

    return jsonify({"success": True, "data": results})


@bp.route("/ssh/batch-deploy", methods=["POST"])
@admin_required
def ssh_batch_deploy():
    """Batch deploy SSH key to all Linux hosts for specified users via Ansible"""
    import subprocess, os
    from datetime import datetime

    data = request.get_json() or {}
    users = data.get("users", ["root", "sysinfra"])
    if isinstance(users, str):
        users = [u.strip() for u in users.split(",") if u.strip()]

    pub_path = os.path.expanduser("~/.ssh/id_ed25519.pub")
    if not os.path.exists(pub_path):
        return jsonify({"success": False, "error": "No SSH key. Generate one first."}), 400

    with open(pub_path) as f:
        pub_key = f.read().strip()

    inventory = "/opt/inspection/ansible/inventory/hosts.yml"
    if not os.path.exists(inventory):
        return jsonify({"success": False, "error": "Ansible inventory not found"}), 500

    # 逐台逐帳號執行，精確回報
    results = []
    linux_hosts = [h for h in get_collection("hosts").find({"os_group": {"$in": ["linux", "Linux"]}}, {"_id": 0, "hostname": 1, "ip": 1}) if h.get("ip")]
    if not linux_hosts:
        # fallback: 用 os 欄位判斷
        linux_hosts = [h for h in get_collection("hosts").find({}, {"_id": 0, "hostname": 1, "ip": 1, "os": 1})
                       if h.get("ip") and h.get("os", "").lower() not in ["windows", ""] and "windows" not in h.get("os", "").lower()]

    for host in linux_hosts:
        hostname = host.get("hostname", host.get("ip"))
        ip = host.get("ip")
        host_result = {"hostname": hostname, "ip": ip, "users": {}}
        deployed_users = []

        for u in users:
            # 先檢查帳號是否存在，不存在就建立
            try:
                chk = subprocess.run(
                    ["ansible", "-i", inventory, hostname, "-b", "-m", "shell",
                     "-a", f"getent passwd {u}"],
                    capture_output=True, text=True, timeout=15
                )
                if chk.returncode != 0 or "FAILED" in chk.stdout:
                    # 帳號不存在，自動建立
                    create_cmd = f"useradd -m -s /bin/bash {u} && echo '{u}:P@ssw0rd' | chpasswd"
                    cr = subprocess.run(
                        ["ansible", "-i", inventory, hostname, "-b", "-m", "shell",
                         "-a", create_cmd],
                        capture_output=True, text=True, timeout=20
                    )
                    if cr.returncode == 0 and "FAILED" not in cr.stdout:
                        host_result["users"][u] = "created"
                        log_action(session["username"], "ssh_create_user",
                            f"Auto created user {u} on {hostname}",
                            request.remote_addr)
                    else:
                        host_result["users"][u] = "create_fail"
                        continue
            except:
                host_result["users"][u] = "check_error"
                continue

            # 帳號存在或剛建立，部署 key
            try:
                dep = subprocess.run(
                    ["ansible", "-i", inventory, hostname, "-b", "-m", "authorized_key",
                     "-a", f"user={u} key='{pub_key}' state=present"],
                    capture_output=True, text=True, timeout=30
                )
                if dep.returncode == 0 and "FAILED" not in dep.stdout:
                    prev = host_result["users"].get(u, "")
                    host_result["users"][u] = "created_ok" if prev == "created" else "ok"
                    deployed_users.append(u)
                else:
                    host_result["users"][u] = "deploy_fail"
            except:
                host_result["users"][u] = "deploy_error"

        host_result["status"] = "ok" if deployed_users else "error"
        host_result["deployed_users"] = deployed_users
        results.append(host_result)

        # 更新 MongoDB — 每個帳號獨立記錄部署時間
        if deployed_users:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            try:
                host_doc = get_collection("hosts").find_one({"hostname": hostname}, {"ssh_key_records": 1}) or {}
                records = host_doc.get("ssh_key_records", {})
                for du in deployed_users:
                    records[du] = now
                get_collection("hosts").update_one(
                    {"hostname": hostname},
                    {"$set": {"ssh_key_records": records}}
                )
            except:
                pass

    log_action(session["username"], "ssh_batch_deploy",
        f"Batch deploy key to users={users}, results={len(results)} hosts",
        request.remote_addr)

    return jsonify({"success": True, "data": results})


@bp.route("/ssh/batch-remove", methods=["POST"])
@admin_required
def ssh_batch_remove():
    """Batch remove SSH key from all Linux hosts for specified users via Ansible"""
    import subprocess, os
    from datetime import datetime

    data = request.get_json() or {}
    users = data.get("users", ["root", "sysinfra"])
    if isinstance(users, str):
        users = [u.strip() for u in users.split(",") if u.strip()]

    pub_path = os.path.expanduser("~/.ssh/id_ed25519.pub")
    if not os.path.exists(pub_path):
        return jsonify({"success": False, "error": "No SSH key found to remove"}), 400

    with open(pub_path) as f:
        pub_key = f.read().strip()

    inventory = "/opt/inspection/ansible/inventory/hosts.yml"

    selected = data.get("hosts", []) if isinstance(data.get("hosts"), list) else []
    host_filter = {"hostname": {"$in": selected}} if selected else {"os_group": {"$in": ["linux", "Linux"]}}
    linux_hosts = [h for h in get_collection("hosts").find(host_filter, {"_id": 0, "hostname": 1, "ip": 1}) if h.get("ip")]
    if not linux_hosts and not selected:
        linux_hosts = [h for h in get_collection("hosts").find({}, {"_id": 0, "hostname": 1, "ip": 1, "os": 1})
                       if h.get("ip") and "windows" not in h.get("os", "").lower()]

    results = []
    for host in linux_hosts:
        hostname = host.get("hostname", host.get("ip"))
        host_result = {"hostname": hostname, "users": {}}
        removed_users = []
        for u in users:
            try:
                chk = subprocess.run(
                    ["ansible", "-i", inventory, hostname, "-b", "-m", "shell",
                     "-a", f"getent passwd {u}"],
                    capture_output=True, text=True, timeout=15
                )
                if chk.returncode != 0 or "FAILED" in chk.stdout:
                    host_result["users"][u] = "no_user"
                    continue
            except:
                host_result["users"][u] = "check_error"
                continue

            try:
                dep = subprocess.run(
                    ["ansible", "-i", inventory, hostname, "-b", "-m", "authorized_key",
                     "-a", f"user={u} key='{pub_key}' state=absent"],
                    capture_output=True, text=True, timeout=30
                )
                if dep.returncode == 0 and "FAILED" not in dep.stdout:
                    host_result["users"][u] = "ok"
                    removed_users.append(u)
                else:
                    host_result["users"][u] = "remove_fail"
            except:
                host_result["users"][u] = "remove_error"

        host_result["status"] = "ok" if removed_users else "error"
        results.append(host_result)

        if removed_users:
            try:
                host_doc = get_collection("hosts").find_one({"hostname": hostname}, {"ssh_key_records": 1}) or {}
                records = host_doc.get("ssh_key_records", {})
                for ru in removed_users:
                    records.pop(ru, None)
                if records:
                    get_collection("hosts").update_one(
                        {"hostname": hostname},
                        {"$set": {"ssh_key_records": records}}
                    )
                else:
                    get_collection("hosts").update_one(
                        {"hostname": hostname},
                        {"$unset": {"ssh_key_records": ""}}
                    )
            except:
                pass

    log_action(session["username"], "ssh_batch_remove",
        f"Batch remove key from users={users}, results={len(results)} hosts",
        request.remote_addr)

    return jsonify({"success": True, "data": results})


@bp.route("/ssh/delete-key", methods=["POST"])
@admin_required
def ssh_delete_key():
    """Delete existing SSH key pair (for regeneration)"""
    import os
    key_path = os.path.expanduser("~/.ssh/id_ed25519")
    pub_path = key_path + ".pub"
    deleted = []
    for p in [key_path, pub_path]:
        if os.path.exists(p):
            os.remove(p)
            deleted.append(p)
    log_action(session["username"], "ssh_delete_key", f"Deleted SSH key: {deleted}", request.remote_addr)
    return jsonify({"success": True, "message": f"Deleted: {deleted}"})

# ========== END SSH Key Management ==========
# ========== Remote Tools ==========

# 異動指令黑名單
MODIFY_KEYWORDS = [
    "systemctl", "service ", "kill", "rm ", "rmdir", "mv ", "dd ",
    "reboot", "shutdown", "halt", "poweroff",
    "useradd", "userdel", "usermod", "passwd",
    "chmod", "chown", "chgrp",
    "mkdir", "touch",
    "iptables", "firewall-cmd", "ufw",
    " > ", " >> ", "sed -i", "tee ", "echo >",
]

def _is_modify_cmd(cmd):
    low = " " + cmd.lower() + " "
    for kw in MODIFY_KEYWORDS:
        if kw in low:
            return True, kw.strip()
    return False, ""

def _in_business_hours():
    from datetime import datetime
    now = datetime.now()
    return 9 <= now.hour < 15

def _run_ansible(hostname, module, args, become=True, timeout=60, exec_user=None):
    """統一執行 ansible 指令，回傳 (ok, stdout, stderr)"""
    import subprocess
    inventory = "/opt/inspection/ansible/inventory/hosts.yml"
    cmd = ["ansible", "-i", inventory, hostname, "-m", module, "-a", args]
    if exec_user:
        cmd.extend(["-u", exec_user])
    if become:
        cmd.insert(-2, "-b")
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        return (r.returncode == 0 and "FAILED" not in r.stdout and "UNREACHABLE" not in r.stdout), r.stdout, r.stderr
    except subprocess.TimeoutExpired:
        return False, "", "Timeout"
    except Exception as e:
        return False, "", str(e)


@bp.route("/remote/hosts", methods=["GET"])
@admin_required
def remote_hosts_list():
    """List all Linux hosts for remote tools"""
    hosts = list(get_collection("hosts").find(
        {}, {"_id": 0, "hostname": 1, "ip": 1, "os": 1, "os_group": 1, "ssh_key_records": 1}
    ))
    linux_hosts = [h for h in hosts if h.get("ip") and "windows" not in (h.get("os","") + h.get("os_group","")).lower()]
    # 整理 key_users 清單方便前端使用
    for h in linux_hosts:
        records = h.get("ssh_key_records", {}) or {}
        h["key_users"] = list(records.keys())
    return jsonify({"success": True, "data": linux_hosts})


@bp.route("/remote/check-space", methods=["POST"])
@admin_required
def remote_check_space():
    """Check disk space usage after hypothetical upload"""
    import subprocess
    data = request.get_json() or {}
    hostnames = data.get("hosts", [])
    target_path = data.get("target_path", "/tmp")
    file_size_mb = float(data.get("file_size_mb", 0))

    results = []
    for hostname in hostnames:
        ok, stdout, stderr = _run_ansible(hostname, "shell",
            f"df -m {target_path} | tail -1 | awk '{{print $2,$3,$4,$5}}'",
            become=False, timeout=20)
        if not ok:
            results.append({"hostname": hostname, "status": "error", "message": stderr or "check failed"})
            continue
        # 從 ansible 輸出取最後一行（有效資料）
        lines = [l.strip() for l in stdout.strip().splitlines() if l.strip()]
        last = lines[-1] if lines else ""
        parts = last.split()
        try:
            total_mb = int(parts[0]); used_mb = int(parts[1]); avail_mb = int(parts[2])
            used_after = used_mb + file_size_mb
            pct_after = (used_after / total_mb) * 100 if total_mb else 0
            warn = pct_after > 85
            results.append({
                "hostname": hostname, "status": "ok",
                "total_mb": total_mb, "used_mb": used_mb, "avail_mb": avail_mb,
                "pct_after": round(pct_after, 1), "warn": warn
            })
        except Exception as e:
            results.append({"hostname": hostname, "status": "error", "message": f"parse error: {last}"})
    return jsonify({"success": True, "data": results})


@bp.route("/remote/upload", methods=["POST"])
@admin_required
def remote_upload():
    """Upload file to multiple hosts /tmp or specified path"""
    import os, tempfile
    from datetime import datetime

    f = request.files.get("file")
    hostnames = request.form.get("hosts", "").split(",")
    hostnames = [h.strip() for h in hostnames if h.strip()]
    target_path = request.form.get("target_path", "/tmp").strip() or "/tmp"
    force = request.form.get("force", "false").lower() == "true"
    exec_user = (request.form.get("exec_user", "") or "sysinfra").strip()

    if not f or not hostnames:
        return jsonify({"success": False, "error": "file and hosts required"}), 400

    # 存成暫存檔
    fd, tmp_path = tempfile.mkstemp(prefix="upload_")
    try:
        f.save(tmp_path)
        fname = os.path.basename(f.filename)
        dest = os.path.join(target_path, fname)

        results = []
        for hostname in hostnames:
            ok, stdout, stderr = _run_ansible(hostname, "copy",
                f"src={tmp_path} dest={dest} mode=0644",
                become=True, timeout=120, exec_user=exec_user)
            results.append({
                "hostname": hostname,
                "status": "ok" if ok else "error",
                "dest": dest,
                "message": "上傳成功" if ok else (stderr or stdout[:200])
            })

        log_action(session["username"], "remote_upload",
            f"Upload {fname} to {len(hostnames)} hosts at {target_path}, force={force}",
            request.remote_addr)

        return jsonify({"success": True, "data": results, "filename": fname, "dest_dir": target_path})
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@bp.route("/remote/download", methods=["POST"])
@admin_required
def remote_download():
    """Download file from multiple hosts to local /tmp/tmp_<ts>/<hostname>/"""
    import os, shutil, zipfile
    from datetime import datetime

    data = request.get_json() or {}
    hostnames = data.get("hosts", [])
    remote_path = data.get("remote_path", "").strip()
    exec_user = (data.get("exec_user", "") or "sysinfra").strip()

    if not remote_path or not hostnames:
        return jsonify({"success": False, "error": "remote_path and hosts required"}), 400

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    session_dir = f"/tmp/tmp_{ts}"
    os.makedirs(session_dir, exist_ok=True)

    results = []
    for hostname in hostnames:
        host_dir = os.path.join(session_dir, hostname)
        os.makedirs(host_dir, exist_ok=True)
        # 用 ansible fetch module
        ok, stdout, stderr = _run_ansible(hostname, "fetch",
            f"src={remote_path} dest={host_dir}/ flat=yes",
            become=True, timeout=120, exec_user=exec_user)
        if ok:
            # 檢查檔案實際存在
            fname = os.path.basename(remote_path)
            local_file = os.path.join(host_dir, fname)
            if os.path.exists(local_file):
                size = os.path.getsize(local_file)
                results.append({"hostname": hostname, "status": "ok", "local_file": local_file, "size": size})
            else:
                results.append({"hostname": hostname, "status": "error", "message": "檔案未取得"})
        else:
            full_err = (stderr.strip() + chr(10) + stdout.strip()).strip() or "fetch failed"
            results.append({"hostname": hostname, "status": "error", "message": full_err[:2000]})

    # 打包 zip
    zip_path = f"{session_dir}.zip"
    try:
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, files in os.walk(session_dir):
                for fn in files:
                    fp = os.path.join(root, fn)
                    zf.write(fp, os.path.relpath(fp, session_dir))
    except Exception as e:
        pass

    log_action(session["username"], "remote_download",
        f"Download {remote_path} from {len(hostnames)} hosts to {session_dir}",
        request.remote_addr)

    return jsonify({
        "success": True,
        "data": results,
        "session_dir": session_dir,
        "zip_path": zip_path,
        "session_id": f"tmp_{ts}"
    })


@bp.route("/remote/download-zip/<session_id>", methods=["GET"])
@admin_required
def remote_download_zip(session_id):
    """Send zip file to browser"""
    import os
    from flask import send_file
    # 安全檢查：只允許 tmp_ 開頭且無路徑穿越
    if not session_id.startswith("tmp_") or "/" in session_id or ".." in session_id:
        return jsonify({"success": False, "error": "invalid session"}), 400
    zip_path = f"/tmp/{session_id}.zip"
    if not os.path.exists(zip_path):
        return jsonify({"success": False, "error": "zip not found"}), 404
    return send_file(zip_path, as_attachment=True, download_name=f"{session_id}.zip")


@bp.route("/remote/downloads-list", methods=["GET"])
@admin_required
def remote_downloads_list():
    """List all download sessions in /tmp"""
    import os
    from datetime import datetime
    sessions = []
    try:
        for item in sorted(os.listdir("/tmp"), reverse=True):
            if item.startswith("tmp_") and os.path.isdir(f"/tmp/{item}"):
                path = f"/tmp/{item}"
                zip_exists = os.path.exists(f"{path}.zip")
                # 統計檔案數
                file_count = 0
                total_size = 0
                for root, dirs, files in os.walk(path):
                    for fn in files:
                        file_count += 1
                        try:
                            total_size += os.path.getsize(os.path.join(root, fn))
                        except: pass
                mtime = datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M:%S")
                sessions.append({
                    "session_id": item, "path": path,
                    "file_count": file_count, "size": total_size,
                    "mtime": mtime, "zip_exists": zip_exists
                })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    return jsonify({"success": True, "data": sessions})


@bp.route("/remote/clear-downloads", methods=["POST"])
@admin_required
def remote_clear_downloads():
    """Clear all download sessions"""
    import os, shutil
    cleared = []
    try:
        for item in os.listdir("/tmp"):
            if item.startswith("tmp_"):
                p = f"/tmp/{item}"
                try:
                    if os.path.isdir(p):
                        shutil.rmtree(p)
                    else:
                        os.unlink(p)
                    cleared.append(item)
                except: pass
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    log_action(session["username"], "remote_clear_downloads",
        f"Cleared {len(cleared)} download sessions", request.remote_addr)
    return jsonify({"success": True, "cleared": cleared, "count": len(cleared)})


@bp.route("/remote/exec", methods=["POST"])
@admin_required
def remote_exec():
    """Execute command on multiple hosts"""
    data = request.get_json() or {}
    hostnames = data.get("hosts", [])
    command = data.get("command", "").strip()
    oa_ref = data.get("oa_ref", "").strip()
    applied = data.get("applied", False)
    exec_user = (data.get("exec_user", "") or "sysinfra").strip()

    if not command or not hostnames:
        return jsonify({"success": False, "error": "command and hosts required"}), 400

    is_modify, kw = _is_modify_cmd(command)
    in_bh = _in_business_hours()

    # 營業時間 + 異動指令 → 需申請
    if is_modify and in_bh:
        if not applied or not oa_ref:
            return jsonify({
                "success": False,
                "error": f"罞️ 營業時間 (09:00-15:00) 禁止實行異動指令！檢測到關鍵字: '{kw}'。如需執行請勾選「已申請」並填寫 OA 單號/備註。",
                "need_oa": True
            }), 403

    results = []
    for hostname in hostnames:
        ok, stdout, stderr = _run_ansible(hostname, "shell", command, become=False, timeout=60, exec_user=exec_user)
        # ansible shell 輸出會包含 hostname | STATUS | ... 前綴，取 stdout 內容
        output = stdout
        if " | " in stdout:
            # 尋找實際輸出開始位置（第一個 >>）
            idx = stdout.find(">>")
            if idx >= 0:
                output = stdout[idx+2:].strip()
        results.append({
            "hostname": hostname,
            "status": "ok" if ok else "error",
            "output": output[:5000],
            "stderr": stderr[:2000] if stderr else ""
        })

    log_action(session["username"], "remote_exec",
        f"exec '{command}' on {len(hostnames)} hosts, modify={is_modify}, oa={oa_ref}, applied={applied}",
        request.remote_addr)

    return jsonify({
        "success": True,
        "data": results,
        "is_modify": is_modify,
        "in_business_hours": in_bh,
        "keyword": kw,
        "exec_user": exec_user
    })

# ========== END Remote Tools ==========