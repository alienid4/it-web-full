"""TWGCB 合規檢查 API"""
from flask import Blueprint, jsonify, request
import subprocess
import json
import os
import glob
from datetime import datetime
from services.mongo_service import get_db

bp = Blueprint("api_twgcb", __name__)

REPORTS_DIR = "/opt/inspection/data/reports"
ANSIBLE_DIR = "/opt/inspection/ansible"
VAULT_PASS = "/opt/inspection/.vault_pass"

@bp.route("/api/twgcb/scan", methods=["POST"])
def trigger_scan():
    """觸發 TWGCB 合規掃描"""
    target = request.json.get("target", "all") if request.is_json else "all"

    limit_arg = "" if target == "all" else f"--limit {target}"
    vault_arg = f"--vault-password-file {VAULT_PASS}" if os.path.exists(VAULT_PASS) else ""
    cmd = f"cd {ANSIBLE_DIR} && ansible-playbook playbooks/twgcb_scan.yml -i inventory/hosts.yml {limit_arg} {vault_arg}"

    try:
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=120)
        if result.returncode != 0:
            return jsonify({"success": False, "error": result.stderr[-500:] if result.stderr else "Playbook failed"}), 500

        saved = _import_results()
        return jsonify({"success": True, "message": f"掃描完成，匯入 {saved} 台主機結果"})
    except subprocess.TimeoutExpired:
        return jsonify({"success": False, "error": "掃描超時 (120s)"}), 504
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/twgcb/import", methods=["POST"])
def import_results():
    """匯入已有的 TWGCB JSON 結果到 MongoDB"""
    count = _import_results()
    return jsonify({"success": True, "message": f"匯入 {count} 台主機結果"})


@bp.route("/api/twgcb/results", methods=["GET"])
def get_results():
    """取得 TWGCB 結果（server-side filter + pagination，支援 500+ 主機）

    Query params:
      os_type: linux|windows|aix|as400
      fail_only: 1/true → 只回有 FAIL 的主機
      ap_owner: 精確比對負責人
      tier: 金|銀|銅
      system: 系統別（regex 模糊比對）
      search: 搜尋 hostname / system_name / ap_owner / custodian（regex）
      limit: 分頁大小（預設 30，最大 100；傳 0 = 不分頁）
      offset: 分頁位移（預設 0）
    回傳：{success, data, total, limit, offset, count}
    """
    db = get_db()
    os_type = request.args.get("os_type", "")
    ap_owner = request.args.get("ap_owner", "").strip()
    tier = request.args.get("tier", "").strip()
    system = request.args.get("system", "").strip()
    search = request.args.get("search", "").strip()
    fail_only = request.args.get("fail_only", "").lower() in ("1", "true", "yes")
    try:
        limit = int(request.args.get("limit", "30"))
    except ValueError:
        limit = 30
    try:
        offset = max(int(request.args.get("offset", "0")), 0)
    except ValueError:
        offset = 0
    if limit < 0:
        limit = 30
    if limit > 100:
        limit = 100  # 硬性上限，防前端誤傳大值拖爆

    # 1. 主機中繼資料篩選 → 在 db.hosts 算出 hostname 白名單
    host_meta_query = {}
    if ap_owner:
        host_meta_query["ap_owner"] = ap_owner
    if tier:
        host_meta_query["tier"] = tier
    if system:
        host_meta_query["system_name"] = {"$regex": system, "$options": "i"}
    if search:
        host_meta_query["$or"] = [
            {"hostname": {"$regex": search, "$options": "i"}},
            {"system_name": {"$regex": search, "$options": "i"}},
            {"ap_owner": {"$regex": search, "$options": "i"}},
            {"custodian": {"$regex": search, "$options": "i"}},
        ]
    matching_hostnames = None
    if host_meta_query:
        matching_hostnames = [h["hostname"] for h in db.hosts.find(host_meta_query, {"hostname": 1, "_id": 0})]
        if not matching_hostnames:
            return jsonify({"success": True, "data": [], "total": 0, "limit": limit, "offset": offset, "count": 0})

    # 2. 組 twgcb_results 查詢
    query = {}
    if os_type == "linux":
        # 變更 #46: 補 "linux" 字面 match，避免 ansible_distribution fallback 成 "Linux" 時主機被漏掉
        query["os"] = {"$regex": "(?i)(rocky|rhel|red hat|centos|debian|ubuntu|suse|oracle linux|linux)"}
    elif os_type == "windows":
        query["os"] = {"$regex": "(?i)windows"}
    elif os_type == "aix":
        query["os"] = {"$regex": "(?i)aix"}
    elif os_type == "as400":
        query["os"] = {"$regex": "(?i)(as.?400|ibm.?i)"}
    if matching_hostnames is not None:
        query["hostname"] = {"$in": matching_hostnames}
    if fail_only:
        query["checks.status"] = "FAIL"

    # 3. 總數（供前端顯示分頁）
    total = db.twgcb_results.count_documents(query)

    # 4. 分頁抓取
    cursor = db.twgcb_results.find(query, {"_id": 0}).sort("hostname", 1).skip(offset)
    if limit > 0:
        cursor = cursor.limit(limit)
    results = list(cursor)

    # 5. 只針對本頁 hostnames 拉中繼 + 例外（500 台 → 只動 30 個，避免全量 scan）
    hostnames_in_page = [r.get("hostname", "") for r in results if r.get("hostname")]
    if hostnames_in_page:
        hosts_info = {h["hostname"]: h for h in db.hosts.find(
            {"hostname": {"$in": hostnames_in_page}},
            {"_id": 0, "hostname": 1, "custodian": 1, "department": 1, "os_group": 1, "system_name": 1, "ap_owner": 1, "tier": 1}
        )}
        exc_map = {}
        for e in db.twgcb_exceptions.find({"hostname": {"$in": hostnames_in_page}}, {"_id": 0}):
            exc_map[(e["hostname"], e["check_id"])] = e
    else:
        hosts_info = {}
        exc_map = {}

    for r in results:
        h = hosts_info.get(r.get("hostname"), {})
        r["custodian"] = h.get("custodian", "")
        r["department"] = h.get("department", "")
        r["system_name"] = h.get("system_name", "")
        r["ap_owner"] = h.get("ap_owner", "")
        r["tier"] = h.get("tier", "")
        hostname = r.get("hostname", "")
        for check in r.get("checks", []):
            cid = check.get("id")
            key = (hostname, cid)
            if key in exc_map:
                check["exception"] = True
                check["exception_reason"] = exc_map[key].get("reason", "")
                check["exception_approved_by"] = exc_map[key].get("approved_by", "")
                check["exception_approved_date"] = exc_map[key].get("approved_date", "")
            else:
                check["exception"] = False

    return jsonify({
        "success": True,
        "data": results,
        "total": total,
        "limit": limit,
        "offset": offset,
        "count": len(results),
    })


@bp.route("/api/twgcb/results/<hostname>", methods=["GET"])
def get_host_result(hostname):
    """取得單台主機的 TWGCB 結果（含例外狀態）"""
    db = get_db()
    result = db.twgcb_results.find_one({"hostname": hostname}, {"_id": 0})
    if not result:
        return jsonify({"success": False, "error": f"找不到 {hostname} 的掃描結果"}), 404

    # 帶入例外資訊
    exceptions = {e["check_id"]: e for e in db.twgcb_exceptions.find({"hostname": hostname}, {"_id": 0})}
    for check in result.get("checks", []):
        cid = check.get("id")
        if cid in exceptions:
            check["exception"] = True
            check["exception_reason"] = exceptions[cid].get("reason", "")
            check["exception_approved_by"] = exceptions[cid].get("approved_by", "")
            check["exception_approved_date"] = exceptions[cid].get("approved_date", "")
        else:
            check["exception"] = False

    return jsonify({"success": True, "data": result})


@bp.route("/api/twgcb/filter-options", methods=["GET"])
def get_filter_options():
    """回傳 distinct 系統別 / AP 負責人 / 級別，供前端 dropdown 一次載滿
    目的：避免分頁後 filter dropdown 只看到當前頁的系統別（500 台規模會踩到）
    """
    db = get_db()
    systems = sorted([s for s in db.hosts.distinct("system_name") if s])
    ap_owners = sorted([a for a in db.hosts.distinct("ap_owner") if a])
    tiers = sorted([t for t in db.hosts.distinct("tier") if t])
    return jsonify({
        "success": True,
        "systems": systems,
        "ap_owners": ap_owners,
        "tiers": tiers,
    })


@bp.route("/api/twgcb/summary", methods=["GET"])
def get_summary():
    """合規總覽摘要"""
    db = get_db()
    results = list(db.twgcb_results.find({}, {"_id": 0}))

    if not results:
        return jsonify({"success": True, "data": {
            "total_hosts": 0, "compliant_hosts": 0, "non_compliant_hosts": 0,
            "compliance_rate": 0, "total_checks": 0,
            "by_level": {}, "by_category": {}, "hosts": []
        }})

    total_hosts = len(results)
    compliant_hosts = 0
    by_level = {}
    by_category = {}
    hosts_summary = []

    for r in results:
        checks = r.get("checks", [])
        total = len(checks)
        passed = sum(1 for c in checks if c.get("status") == "PASS")
        failed = total - passed
        rate = round(passed / total * 100, 1) if total > 0 else 0

        if failed == 0:
            compliant_hosts += 1

        hosts_summary.append({
            "hostname": r.get("hostname"),
            "os": r.get("os", ""),
            "scan_time": r.get("scan_time", ""),
            "total": total,
            "passed": passed,
            "failed": failed,
            "compliance_rate": rate
        })

        for c in checks:
            level = c.get("level", "?")
            cat = c.get("category", "?")
            status = c.get("status", "FAIL")

            by_level.setdefault(level, {"total": 0, "passed": 0})
            by_level[level]["total"] += 1
            if status == "PASS":
                by_level[level]["passed"] += 1

            by_category.setdefault(cat, {"total": 0, "passed": 0})
            by_category[cat]["total"] += 1
            if status == "PASS":
                by_category[cat]["passed"] += 1

    for k in by_level:
        t = by_level[k]["total"]
        by_level[k]["rate"] = round(by_level[k]["passed"] / t * 100, 1) if t > 0 else 0
    for k in by_category:
        t = by_category[k]["total"]
        by_category[k]["rate"] = round(by_category[k]["passed"] / t * 100, 1) if t > 0 else 0

    total_checks = sum(len(r.get("checks", [])) for r in results)
    total_passed = sum(sum(1 for c in r.get("checks", []) if c.get("status") == "PASS") for r in results)
    overall_rate = round(total_passed / total_checks * 100, 1) if total_checks > 0 else 0

    return jsonify({"success": True, "data": {
        "total_hosts": total_hosts,
        "compliant_hosts": compliant_hosts,
        "non_compliant_hosts": total_hosts - compliant_hosts,
        "compliance_rate": overall_rate,
        "total_checks": total_checks,
        "total_passed": total_passed,
        "by_level": by_level,
        "by_category": by_category,
        "hosts": sorted(hosts_summary, key=lambda x: x["compliance_rate"])
    }})


@bp.route("/api/twgcb/check/<check_id>", methods=["GET"])
def get_check_detail(check_id):
    """依檢查項看所有主機狀態"""
    db = get_db()
    results = list(db.twgcb_results.find({}, {"_id": 0}))

    hosts_pass = []
    hosts_fail = []
    check_info = None

    for r in results:
        for c in r.get("checks", []):
            if c.get("id") == check_id:
                if not check_info:
                    check_info = {k: c[k] for k in ["id", "category", "name", "level", "expected"]}
                entry = {
                    "hostname": r.get("hostname"),
                    "os": r.get("os", ""),
                    "actual": c.get("actual", ""),
                    "detail": c.get("detail", ""),
                    "remediation": c.get("remediation", "")
                }
                if c.get("status") == "PASS":
                    hosts_pass.append(entry)
                else:
                    hosts_fail.append(entry)

    if not check_info:
        return jsonify({"success": False, "error": f"找不到檢查項 {check_id}"}), 404

    return jsonify({"success": True, "data": {
        "check": check_info,
        "pass_count": len(hosts_pass),
        "fail_count": len(hosts_fail),
        "hosts_pass": hosts_pass,
        "hosts_fail": hosts_fail
    }})


@bp.route("/api/twgcb/remediate", methods=["POST"])
def remediate():
    """執行修復指令"""
    if not request.is_json:
        return jsonify({"success": False, "error": "需要 JSON"}), 400

    hostname = request.json.get("hostname")
    check_id = request.json.get("check_id")
    remediation = request.json.get("remediation")

    if not all([hostname, check_id, remediation]):
        return jsonify({"success": False, "error": "缺少 hostname/check_id/remediation"}), 400

    # 安全檢查：禁止危險指令
    dangerous = ["rm -rf", "mkfs", "dd if=", "> /dev/", "shutdown", "reboot", "init 0", "init 6"]
    for d in dangerous:
        if d in remediation:
            return jsonify({"success": False, "error": f"拒絕執行危險指令: {d}"}), 403

    # 透過 Ansible 在目標主機執行修復
    # -b (become): TWGCB 修復多為 systemctl/rpm/sysctl，需要 root 權限，必加 sudo
    vault_arg = f"--vault-password-file {VAULT_PASS}" if os.path.exists(VAULT_PASS) else ""
    cmd = f'cd {ANSIBLE_DIR} && ansible {hostname} -i inventory/hosts.yml -b -m shell -a "{remediation}" {vault_arg}'

    try:
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=30)
        output = result.stdout + result.stderr
        # Ansible shell module 成功輸出格式: "<host> | CHANGED | rc=0 >>" 或 "| SUCCESS |"
        # 失敗時: "| FAILED |" / "| UNREACHABLE!"
        success = (result.returncode == 0
                   and "| FAILED" not in result.stdout
                   and "UNREACHABLE" not in result.stdout)

        return jsonify({
            "success": success,
            "hostname": hostname,
            "check_id": check_id,
            "output": output[-1000:] if output else "無輸出",
            "message": f"{'修復成功' if success else '修復失敗'}: {check_id} on {hostname}"
        })
    except subprocess.TimeoutExpired:
        return jsonify({"success": False, "error": "修復超時 (30s)"}), 504
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


def _import_results():
    """將 JSON 檔匯入 MongoDB"""
    db = get_db()
    pattern = os.path.join(REPORTS_DIR, "twgcb_*.json")
    files = glob.glob(pattern)
    count = 0
    for f in files:
        try:
            with open(f) as fh:
                data = json.load(fh)
            hostname = data.get("hostname")
            if not hostname:
                continue
            data["imported_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            db.twgcb_results.update_one(
                {"hostname": hostname},
                {"$set": data},
                upsert=True
            )
            count += 1
        except Exception:
            continue
    return count
"""追加到 api_twgcb.py 末尾的設定管理 + 報表 API"""


# ========== 設定管理 API ==========

@bp.route("/api/twgcb/config", methods=["GET"])
def get_config():
    """取得所有 TWGCB 檢查項設定"""
    db = get_db()
    configs = list(db.twgcb_config.find({}, {"_id": 0}).sort("check_id", 1))
    return jsonify({"success": True, "data": configs, "count": len(configs)})


@bp.route("/api/twgcb/config/<check_id>", methods=["PUT"])
def update_config(check_id):
    """更新單項檢查設定（啟用/停用、閾值、例外主機）"""
    if not request.is_json:
        return jsonify({"success": False, "error": "需要 JSON"}), 400

    db = get_db()
    existing = db.twgcb_config.find_one({"check_id": check_id})
    if not existing:
        return jsonify({"success": False, "error": f"找不到 {check_id}"}), 404

    allowed_fields = ["enabled", "threshold", "expected", "exception_hosts", "remediation", "description"]
    update = {}
    for f in allowed_fields:
        if f in request.json:
            update[f] = request.json[f]
    update["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    db.twgcb_config.update_one({"check_id": check_id}, {"$set": update})
    return jsonify({"success": True, "message": f"{check_id} 設定已更新"})


# ========== 報表設定 API ==========

@bp.route("/api/twgcb/report-config", methods=["GET"])
def get_report_config():
    """取得報表設定"""
    db = get_db()
    config = db.twgcb_report_config.find_one({"key": "twgcb_report"}, {"_id": 0})
    if not config:
        config = {"company_name": "Example Corp", "department": "", "handler": "", "section_chief": "", "director": ""}
    return jsonify({"success": True, "data": config})


@bp.route("/api/twgcb/report-config", methods=["PUT"])
def update_report_config():
    """更新報表設定"""
    if not request.is_json:
        return jsonify({"success": False, "error": "需要 JSON"}), 400

    db = get_db()
    allowed = ["company_name", "department", "handler", "section_chief", "director", "logo_text"]
    update = {f: request.json[f] for f in allowed if f in request.json}
    update["key"] = "twgcb_report"

    db.twgcb_report_config.update_one(
        {"key": "twgcb_report"}, {"$set": update}, upsert=True
    )
    return jsonify({"success": True, "message": "報表設定已更新"})


# ========== 報表產生 API ==========

@bp.route("/api/twgcb/report", methods=["GET"])
def generate_report():
    """產生 TWGCB 合規報表資料（燈號矩陣）"""
    db = get_db()

    # 取得設定
    configs = {c["check_id"]: c for c in db.twgcb_config.find({}, {"_id": 0})}
    enabled_checks = [c for c in configs.values() if c.get("enabled", True)]
    enabled_ids = [c["check_id"] for c in enabled_checks]

    # 取得報表設定
    report_cfg = db.twgcb_report_config.find_one({"key": "twgcb_report"}, {"_id": 0}) or {}

    # 取得掃描結果
    results = list(db.twgcb_results.find({}, {"_id": 0}))

    # 建立燈號矩陣
    matrix = []
    for r in sorted(results, key=lambda x: x.get("hostname", "")):
        hostname = r.get("hostname", "")
        checks = r.get("checks", [])
        row = {"hostname": hostname, "os": r.get("os", ""), "checks": {}}
        passed = 0
        total = 0

        for c in checks:
            cid = c.get("id")
            if cid not in enabled_ids:
                continue

            total += 1
            cfg = configs.get(cid, {})
            exception_hosts = cfg.get("exception_hosts", [])

            if hostname in exception_hosts:
                row["checks"][cid] = {"status": "exception", "actual": c.get("actual", ""), "detail": c.get("detail", "")}
            elif c.get("status") == "PASS":
                row["checks"][cid] = {"status": "pass", "actual": c.get("actual", ""), "detail": c.get("detail", "")}
                passed += 1
            else:
                row["checks"][cid] = {"status": "fail", "actual": c.get("actual", ""), "detail": c.get("detail", ""),
                                       "remediation": c.get("remediation", "")}

        row["total"] = total
        row["passed"] = passed
        row["failed"] = total - passed
        row["rate"] = round(passed / total * 100, 1) if total > 0 else 0
        matrix.append(row)

    # 分類欄位
    categories = {}
    for c in enabled_checks:
        cat = c.get("category", "其他")
        if cat not in categories:
            categories[cat] = []
        categories[cat].append({
            "check_id": c["check_id"],
            "name": c.get("name", c.get("check_id", "")),
            "level": c.get("level", ""),
            "threshold": c.get("threshold", c.get("expected", ""))
        })

    return jsonify({"success": True, "data": {
        "report_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "company_name": report_cfg.get("company_name", "Example Corp"),
        "department": report_cfg.get("department", ""),
        "handler": report_cfg.get("handler", ""),
        "section_chief": report_cfg.get("section_chief", ""),
        "director": report_cfg.get("director", ""),
        "categories": categories,
        "enabled_checks": [{
            "check_id": c["check_id"], "name": c.get("name", c.get("check_id", "")),
            "category": c.get("category", "其他"), "level": c.get("level", "")
        } for c in enabled_checks],
        "matrix": matrix,
        "total_hosts": len(matrix),
        "total_checks": len(enabled_ids)
    }})


# ========== 例外管理 API ==========

@bp.route("/api/twgcb/exceptions", methods=["GET"])
def get_exceptions():
    """取得所有 TWGCB 例外"""
    db = get_db()
    hostname = request.args.get("hostname")
    query = {"hostname": hostname} if hostname else {}
    exceptions = list(db.twgcb_exceptions.find(query, {"_id": 0}))
    return jsonify({"success": True, "data": exceptions, "count": len(exceptions)})


@bp.route("/api/twgcb/exceptions", methods=["POST"])
def add_exception():
    """新增例外"""
    if not request.is_json:
        return jsonify({"success": False, "error": "需要 JSON"}), 400

    check_id = request.json.get("check_id", "").strip()
    hostname = request.json.get("hostname", "").strip()
    reason = request.json.get("reason", "").strip()
    approved_by = request.json.get("approved_by", "").strip()

    if not all([check_id, hostname, reason]):
        return jsonify({"success": False, "error": "缺少 check_id / hostname / reason"}), 400

    db = get_db()
    # 檢查是否已存在
    existing = db.twgcb_exceptions.find_one({"check_id": check_id, "hostname": hostname})
    if existing:
        # 更新
        db.twgcb_exceptions.update_one(
            {"check_id": check_id, "hostname": hostname},
            {"$set": {
                "reason": reason,
                "approved_by": approved_by,
                "approved_date": request.json.get("approved_date", datetime.now().strftime("%Y-%m-%d")),
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }}
        )
        return jsonify({"success": True, "message": f"{hostname} / {check_id} 例外已更新"})

    # 新增
    doc = {
        "check_id": check_id,
        "hostname": hostname,
        "reason": reason,
        "approved_by": approved_by,
        "approved_date": request.json.get("approved_date", datetime.now().strftime("%Y-%m-%d")),
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    db.twgcb_exceptions.insert_one(doc)
    return jsonify({"success": True, "message": f"{hostname} / {check_id} 已標記為例外"})


@bp.route("/api/twgcb/exceptions", methods=["DELETE"])
def delete_exception():
    """移除例外"""
    if not request.is_json:
        return jsonify({"success": False, "error": "需要 JSON"}), 400

    check_id = request.json.get("check_id", "").strip()
    hostname = request.json.get("hostname", "").strip()

    if not all([check_id, hostname]):
        return jsonify({"success": False, "error": "缺少 check_id / hostname"}), 400

    db = get_db()
    result = db.twgcb_exceptions.delete_one({"check_id": check_id, "hostname": hostname})
    if result.deleted_count == 0:
        return jsonify({"success": False, "error": "找不到該例外"}), 404
    return jsonify({"success": True, "message": f"{hostname} / {check_id} 例外已移除"})


# ========== Excel 匯出 API ==========

@bp.route("/api/twgcb/export", methods=["GET"])
def export_excel():
    """匯出 TWGCB 完整矩陣 Excel"""
    import io
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        return jsonify({"success": False, "error": "需要 openpyxl：pip3 install openpyxl"}), 500

    db = get_db()
    os_type = request.args.get("os_type", "")
    query = {}
    if os_type == "linux":
        # 變更 #46: 補 "linux" 字面 match，避免 ansible_distribution fallback 成 "Linux" 時主機被漏掉
        query["os"] = {"$regex": "(?i)(rocky|rhel|red hat|centos|debian|ubuntu|suse|oracle linux|linux)"}
    elif os_type == "windows":
        query["os"] = {"$regex": "(?i)windows"}
    elif os_type == "aix":
        query["os"] = {"$regex": "(?i)aix"}
    elif os_type == "as400":
        query["os"] = {"$regex": "(?i)(as.?400|ibm.?i)"}

    results = list(db.twgcb_results.find(query, {"_id": 0}).sort("hostname", 1))
    if not results:
        return jsonify({"success": False, "error": "無資料可匯出"}), 404

    # 例外
    all_exc = {(e["hostname"],e["check_id"]): e for e in db.twgcb_exceptions.find({}, {"_id": 0})}
    # 主機資訊
    hosts_info = {h["hostname"]: h for h in db.hosts.find({}, {"_id": 0})}

    # 收集檢查項
    check_map = {}
    check_order = []
    for r in results:
        for c in r.get("checks", []):
            if c["id"] not in check_map:
                check_map[c["id"]] = c
                check_order.append(c["id"])

    hostnames = [r["hostname"] for r in results]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TWGCB合規矩陣"

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    orange_fill = PatternFill("solid", fgColor="FFE0B2")
    header_fill = PatternFill("solid", fgColor="E8F5E9")
    bold = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Row 1: 系統別
    ws.cell(1, 1, "系統別").font = bold
    for i, h in enumerate(hostnames):
        hi = hosts_info.get(h, {})
        ws.cell(1, 5+i, hi.get("system_name", "-"))

    # Row 2: AP負責人
    ws.cell(2, 1, "AP負責人").font = bold
    for i, h in enumerate(hostnames):
        hi = hosts_info.get(h, {})
        ws.cell(2, 5+i, hi.get("ap_owner", hi.get("custodian", "-")))

    # Row 3: 級別
    ws.cell(3, 1, "級別").font = bold
    for i, h in enumerate(hostnames):
        hi = hosts_info.get(h, {})
        ws.cell(3, 5+i, hi.get("tier", "-"))

    # Row 4: Header
    headers = ["#", "分類", "編號", "說明"] + hostnames + ["例外說明"]
    for j, hdr in enumerate(headers):
        cell = ws.cell(4, j+1, hdr)
        cell.font = bold
        cell.fill = header_fill
        cell.border = thin_border

    # Data rows
    host_check_map = {}
    for r in results:
        host_check_map[r["hostname"]] = {c["id"]: c for c in r.get("checks", [])}

    for idx, cid in enumerate(check_order):
        ck = check_map[cid]
        row = 5 + idx
        ws.cell(row, 1, idx+1).border = thin_border
        ws.cell(row, 2, ck.get("category","")).border = thin_border
        ws.cell(row, 3, ck["id"]).border = thin_border
        ws.cell(row, 4, ck["name"]).border = thin_border

        reasons = []
        for i, h in enumerate(hostnames):
            cell = ws.cell(row, 5+i)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            c = host_check_map.get(h, {}).get(cid)
            if not c:
                cell.value = "-"
            elif c.get("status") == "PASS":
                cell.value = "PASS"
                cell.fill = green_fill
            elif (h, cid) in all_exc:
                cell.value = "例外"
                cell.fill = orange_fill
                reasons.append(h + "：" + all_exc[(h,cid)].get("reason",""))
            else:
                cell.value = "FAIL"
                cell.fill = red_fill

        reason_cell = ws.cell(row, 5+len(hostnames))
        reason_cell.value = "; ".join(reasons)
        reason_cell.border = thin_border

    # Auto width
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    os_label = os_type or "all"
    filename = "TWGCB_" + os_label + "_" + datetime.now().strftime("%Y%m%d_%H%M") + ".xlsx"
    return send_file(buf, download_name=filename, as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ========== Excel 匯入 API (AS400 等手動匯入) ==========

@bp.route("/api/twgcb/import-excel", methods=["POST"])
def import_excel():
    """從 Excel 匯入 TWGCB 結果 (AS400/手動設備)"""
    try:
        import openpyxl
    except ImportError:
        return jsonify({"success": False, "error": "需要 openpyxl"}), 500

    if "file" not in request.files:
        return jsonify({"success": False, "error": "請上傳 Excel 檔案"}), 400

    f = request.files["file"]
    wb = openpyxl.load_workbook(f)
    ws = wb.active

    db = get_db()
    imported = {}  # hostname -> checks list
    rows_read = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        hostname = str(row[0]).strip()
        check_id = str(row[1]).strip() if row[1] else ""
        status = str(row[2]).strip().upper() if row[2] else "FAIL"
        actual = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        name = str(row[4]).strip() if len(row) > 4 and row[4] else check_id
        category = str(row[5]).strip() if len(row) > 5 and row[5] else "匯入項目"
        level = str(row[6]).strip() if len(row) > 6 and row[6] else "A"

        if not check_id:
            continue

        if hostname not in imported:
            imported[hostname] = []

        imported[hostname].append({
            "id": check_id,
            "category": category,
            "name": name,
            "level": level,
            "expected": "",
            "actual": actual,
            "status": "PASS" if status == "PASS" else "FAIL",
            "detail": actual,
            "remediation": "",
            "gcb_ref": check_id
        })
        rows_read += 1

    # 寫入 MongoDB
    count = 0
    for hostname, checks in imported.items():
        doc = {
            "hostname": hostname,
            "scan_time": datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ"),
            "os": "AS/400" if "as400" in hostname.lower() or "as400" in (request.form.get("os_type","")).lower() else "Manual Import",
            "checks": checks,
            "imported_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "import_source": "excel"
        }
        db.twgcb_results.update_one({"hostname": hostname}, {"$set": doc}, upsert=True)
        count += 1

    return jsonify({"success": True, "message": f"匯入 {count} 台主機，共 {rows_read} 筆檢查項"})

@bp.route("/api/twgcb/stats", methods=["GET"])
def get_twgcb_stats():
    """TWGCB 統計: 總覽 / 按主機 / 按分類 / FAIL 項排行"""
    db = get_db()
    results = list(db.twgcb_results.find({}, {"_id": 0}))
    # 主機統計
    by_host = []
    total_pass, total_fail = 0, 0
    cat_stats = {}  # category → {pass, fail}
    fail_counter = {}  # (id, name) → count

    for r in results:
        h = r.get("hostname", "")
        os_name = r.get("os", "")
        checks = r.get("checks", [])
        p = sum(1 for c in checks if c.get("status") == "PASS")
        f = sum(1 for c in checks if c.get("status") == "FAIL")
        total = p + f
        by_host.append({
            "hostname": h,
            "os": os_name,
            "pass": p, "fail": f, "total": total,
            "rate": round(p / total * 100, 1) if total else 0,
        })
        total_pass += p
        total_fail += f
        for c in checks:
            cat = c.get("category", "其他")
            if cat not in cat_stats:
                cat_stats[cat] = {"pass": 0, "fail": 0}
            if c.get("status") == "PASS":
                cat_stats[cat]["pass"] += 1
            elif c.get("status") == "FAIL":
                cat_stats[cat]["fail"] += 1
                key = (c.get("id"), c.get("name", ""))
                fail_counter[key] = fail_counter.get(key, 0) + 1

    # 分類整理
    by_category = []
    for cat, s in cat_stats.items():
        t = s["pass"] + s["fail"]
        by_category.append({
            "category": cat,
            "pass": s["pass"], "fail": s["fail"], "total": t,
            "rate": round(s["pass"] / t * 100, 1) if t else 0,
        })
    by_category.sort(key=lambda x: x["rate"])

    # FAIL 熱點 Top 10
    top_fails = sorted(fail_counter.items(), key=lambda x: -x[1])[:10]
    top_fails = [{"id": k[0], "name": k[1], "count": v} for k, v in top_fails]

    overall = {
        "total": total_pass + total_fail,
        "pass": total_pass,
        "fail": total_fail,
        "rate": round(total_pass / (total_pass + total_fail) * 100, 1) if (total_pass + total_fail) else 0,
        "host_count": len(results),
    }

    return jsonify({
        "success": True,
        "overall": overall,
        "by_host": by_host,
        "by_category": by_category,
        "top_fails": top_fails,
    })
